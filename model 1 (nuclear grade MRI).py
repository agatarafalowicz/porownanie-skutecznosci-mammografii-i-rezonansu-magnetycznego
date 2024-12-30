import os
from collections import defaultdict

import pandas as pd
import numpy as np
import cv2
import torch
import torch.nn as nn
from torch.utils.data import Dataset, DataLoader, WeightedRandomSampler, ConcatDataset
from transformers import ViTModel
from sklearn.model_selection import train_test_split
from sklearn.metrics import classification_report
from sklearn.utils.class_weight import compute_class_weight
from tqdm import tqdm
import albumentations as A
from albumentations.pytorch import ToTensorV2
import matplotlib.pyplot as plt
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import pydicom


class MRIDataProcessor:
    """
    Klasa do przetwarzania danych MRI.
    """

    def __init__(self, base_path, roi_excel_path, clinical_excel_path, processed_folder, output_excel_path):
        """
        Inicjalizuje obiekt MRIDataProcessor.

        Parametry:
            base_path (str): Ścieżka do folderu bazowego z danymi DICOM.
            roi_excel_path (str): Ścieżka do pliku Excel z informacjami o ROI (regionach zainteresowania).
            clinical_excel_path (str): Ścieżka do pliku Excel z danymi klinicznymi.
            processed_folder (str): Ścieżka do folderu, w którym będą zapisywane przetworzone obrazy.
            output_excel_path (str): Ścieżka do pliku Excel, w którym będą zapisywane podsumowania.

        Zwraca:
            None
        """
        self.base_path = base_path
        self.roi_excel_path = roi_excel_path
        self.clinical_excel_path = clinical_excel_path
        self.processed_folder = processed_folder
        self.output_excel_path = output_excel_path
        os.makedirs(self.processed_folder, exist_ok=True)

    def load_roi_and_slices_from_excel(self, patient_id):
        """
        Wczytuje z pliku Excel współrzędne ROI oraz indeksy warstw dla danego pacjenta.

        Parametry:
            patient_id (str): Identyfikator pacjenta.

        Zwraca:
            tuple: Krotka zawierająca (start_row, end_row, start_column, end_column, start_slice, end_slice).

        Wyjątki:
            ValueError: Jeśli brak danych ROI dla danego pacjenta.
        """
        df = pd.read_excel(self.roi_excel_path)
        roi_data = df[df["Patient ID"] == patient_id]
        if roi_data.empty:
            raise ValueError(f"Brak danych ROI dla pacjenta {patient_id}")
        return (
            roi_data["Start Row"].values[0] - 10,
            roi_data["End Row"].values[0] + 10,
            roi_data["Start Column"].values[0] - 10,
            roi_data["End Column"].values[0] + 10,
            roi_data["Start Slice"].values[0],
            roi_data["End Slice"].values[0],
        )

    def load_clinical_labels(self, patient_id):
        """
        Wczytuje dane kliniczne (tumor grade i molecular subtype) z pliku Excel dla danego pacjenta.

        Parametry:
            patient_id (str): Identyfikator pacjenta.

        Zwraca:
            tuple or (None, None): (grade, subtype) lub (None, None), jeśli dane nie są dostępne.
        """
        df = pd.read_excel(self.clinical_excel_path, header=None)
        df.columns = [' '.join(map(str, col)).replace("\n", " ").strip() for col in zip(*df.iloc[:3].values)]
        df = df[3:]
        patient_id_col = [col for col in df.columns if "Breast_MRI" in str(df[col].values)]
        tumor_grade_col = [col for col in df.columns if "Tumor Grade(N)" in col]
        mol_subtype_col = [col for col in df.columns if "Mol Subtype" in col]
        clinical_data = df[df[patient_id_col[0]] == patient_id]
        if clinical_data.empty:
            return None, None
        return clinical_data[tumor_grade_col[0]].values[0], clinical_data[mol_subtype_col[0]].values[0]

    def process_and_save_data(self):
        """
        Przetwarza dane DICOM, wycina obszar ROI, stosuje przetwarzanie wstępne (median blur, CLAHE, normalizacja),
        zapisuje wynik do pliku PNG oraz tworzy plik CSV z etykietami i plik Excel z podsumowaniem.

        Zwraca:
            None
        """
        data_summary = []
        labels = []
        for patient_id in tqdm(os.listdir(self.base_path), desc="Przetwarzanie pacjentów"):
            try:
                image_filename = f"{patient_id}.png"
                processed_path = os.path.join(self.processed_folder, image_filename)

                if os.path.exists(processed_path):
                    print(f"Dane dla pacjenta {patient_id} już istnieją. Pomijanie...")
                    continue

                start_row, end_row, start_column, end_column, start_slice, end_slice = self.load_roi_and_slices_from_excel(
                    patient_id
                )
                dicom_folder = self.find_matching_dicom_folder(patient_id)
                dicom_files = sorted(
                    [os.path.join(dicom_folder, f) for f in os.listdir(dicom_folder) if f.endswith(".dcm")]
                )
                slice_range = dicom_files[start_slice:end_slice + 1]
                if not slice_range:
                    print(f"Brak plików DICOM w zakresie dla pacjenta {patient_id}. Pomijanie...")
                    continue
                middle_index = len(slice_range) // 2
                dicom_data = pydicom.dcmread(slice_range[middle_index])
                image = dicom_data.pixel_array
                roi_image = image[start_row:end_row, start_column:end_column]
                roi_image = cv2.medianBlur(roi_image, 3)
                clahe = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(8, 8))
                normalized_image = cv2.normalize(roi_image, None, 0, 255, cv2.NORM_MINMAX).astype(np.uint8)
                processed_image = clahe.apply(normalized_image)
                processed_image = cv2.resize(processed_image, (224, 224))
                cv2.imwrite(processed_path, processed_image)
                grade, subtype = self.load_clinical_labels(patient_id)
                if grade is not None and subtype is not None:
                    labels.append({'filename': image_filename, 'subtype': subtype, 'grade': grade})
                    data_summary.append(
                        {"Patient ID": patient_id, "Mol Subtype": subtype, "Tumor Grade (Nuclear)": grade}
                    )
            except Exception as e:
                print(f"Nie udało się przetworzyć danych pacjenta {patient_id}: {e}")

        labels_df = pd.DataFrame(labels)
        labels_csv_path = os.path.join(self.processed_folder, "labels.csv")
        labels_df.to_csv(labels_csv_path, index=False)
        df_summary = pd.DataFrame(data_summary)
        with pd.ExcelWriter(self.output_excel_path, engine="openpyxl") as writer:
            df_summary.to_excel(writer, sheet_name="Processed_Data", index=False)
        print("Podsumowanie zapisano do pliku Excel.")

    def find_matching_dicom_folder(self, patient_id):
        """
        Znajduje folder zawierający pliki DICOM dla pacjenta.

        Parametry:
            patient_id (str): Identyfikator pacjenta.

        Zwraca:
            str: Ścieżka do folderu z plikami DICOM.

        Wyjątki:
            FileNotFoundError: Jeśli folder pacjenta nie istnieje lub nie znaleziono plików DICOM.
        """
        patient_path = os.path.join(self.base_path, patient_id)
        if not os.path.exists(patient_path):
            raise FileNotFoundError(f"Folder pacjenta {patient_id} nie istnieje: {patient_path}")

        dicom_folders = []
        for root, dirs, files in os.walk(patient_path):
            if any(file.endswith('.dcm') for file in files):
                dicom_folders.append(root)

        if not dicom_folders:
            raise FileNotFoundError(f"Nie znaleziono plików DICOM dla pacjenta {patient_id}.")
        return dicom_folders[0]

    def calculate_mean_std(self, images):
        """
        Oblicza średnią (mean) oraz odchylenie standardowe (std) dla zestawu obrazów.

        Parametry:
            images (list or numpy.ndarray): Lista lub tablica obrazów w formacie [N, W, H, C].

        Zwraca:
            tuple: (mean, std) - średnia i odchylenie standardowe dla każdego kanału.
        """
        all_pixels = np.concatenate([img.flatten().reshape(-1, 3) for img in images], axis=0)
        mean = np.mean(all_pixels, axis=0) / 255.0
        std = np.std(all_pixels, axis=0) / 255.0
        return mean, std

    def load_filtered_data(self):
        """
        Wczytuje obrazy i etykiety z pliku CSV, filtruje niepoprawne wartości etykiet i zwraca
        obrazy oraz etykiety do dalszego przetwarzania.

        Zwraca:
            tuple: (images, labels, nuclear_grade_counts)
                   images (numpy.ndarray) - Tablica obrazów.
                   labels (numpy.ndarray) - Tablica etykiet (po zmapowaniu).
                   nuclear_grade_counts (dict) - Słownik z liczbą próbek w każdej klasie.

        Wyjątki:
            FileNotFoundError: Jeśli plik labels.csv nie istnieje.
        """
        images, labels = [], []
        nuclear_grade_counts = {}
        labels_csv_path = os.path.join(self.processed_folder, "labels.csv")

        if not os.path.exists(labels_csv_path):
            raise FileNotFoundError(f"Plik z etykietami nie został znaleziony: {labels_csv_path}")

        labels_df = pd.read_csv(labels_csv_path)
        for index, row in tqdm(labels_df.iterrows(), total=labels_df.shape[0],
                               desc="Wczytywanie przetworzonych obrazów"):
            filename = row['filename']
            image_path = os.path.join(self.processed_folder, filename)
            label = row['grade']
            image = cv2.imread(image_path, cv2.IMREAD_GRAYSCALE)
            if image is None:
                print(f"Nie można wczytać obrazu: {image_path}")
                continue
            image = np.dstack([image] * 3).astype(np.float32) / 255.0
            nuclear_grade = label
            if pd.isna(nuclear_grade):
                print(f"Pomijanie obrazu {filename}, ponieważ Nuclear Grade jest NaN.")
                continue
            try:
                nuclear_grade = int(nuclear_grade)
            except ValueError:
                print(f"Nieprawidłowa wartość Nuclear Grade dla obrazu {filename}. Pomijanie...")
                continue
            images.append(image)
            labels.append(nuclear_grade)
            nuclear_grade_counts[nuclear_grade] = nuclear_grade_counts.get(nuclear_grade, 0) + 1

        label_mapping = {1: 0, 2: 1, 3: 2}
        labels = np.array([label_mapping.get(label, -1) for label in labels])
        valid_indices = labels != -1
        images = np.array(images)[valid_indices]
        labels = labels[valid_indices]
        return images, labels, nuclear_grade_counts

    def save_split_data(self, images, labels, folder, labels_csv_path):
        """
        Zapisuje obrazy i odpowiadające im etykiety do określonego folderu oraz generuje plik CSV z etykietami.

        Parametry:
            images (numpy.ndarray): Tablica obrazów w formacie [N, W, H, C].
            labels (numpy.ndarray): Tablica etykiet w formacie [N].
            folder (str): Ścieżka do folderu, w którym zostaną zapisane obrazy.
            labels_csv_path (str): Ścieżka do pliku CSV z etykietami.

        Zwraca:
            None
        """
        os.makedirs(folder, exist_ok=True)
        labels_list = []
        for idx, (img, lbl) in enumerate(zip(images, labels)):
            filename = f"{idx}.png"
            img_uint8 = (img * 255).astype(np.uint8)
            image_path = os.path.join(folder, filename)
            cv2.imwrite(image_path, img_uint8)
            labels_list.append({'filename': filename, 'grade': lbl})
        labels_df = pd.DataFrame(labels_list)
        labels_df.to_csv(labels_csv_path, index=False)
        print(f"Saved {len(images)} images to {folder} with labels at {labels_csv_path}.")

    @staticmethod
    def ensure_train_test_split(images, labels, test_size=0.1, val_size=0.1, random_state=42):
        """
        Dzieli zbiór na podzbiory treningowy, walidacyjny i testowy z zachowaniem proporcji klas.

        Parametry:
            images (numpy.ndarray): Tablica obrazów.
            labels (numpy.ndarray): Tablica etykiet.
            test_size (float): Rozmiar zbioru testowego (proporcja).
            val_size (float): Rozmiar zbioru walidacyjnego (proporcja).
            random_state (int): Ziarno losowe do powtarzalności wyników.

        Zwraca:
            tuple: (x_train, x_val, x_test, y_train, y_val, y_test)
        """
        x_train, x_temp, y_train, y_temp = train_test_split(
            images, labels, test_size=test_size + val_size, random_state=random_state, stratify=labels
        )
        val_ratio = val_size / (test_size + val_size)
        x_val, x_test, y_val, y_test = train_test_split(
            x_temp, y_temp, test_size=1 - val_ratio, random_state=random_state, stratify=y_temp
        )
        return x_train, x_val, x_test, y_train, y_val, y_test


class BreastCancerDataset(Dataset):
    """
    Klasa Dataset dla obrazów raka piersi, kompatybilna z PyTorch.
    """

    def __init__(self, images_folder, labels_csv, augment=False, mean=None, std=None):
        """
        Inicjalizuje dataset, wczytując ścieżki do obrazów i etykiet z pliku CSV.

        Parametry:
            images_folder (str): Ścieżka do folderu z obrazami.
            labels_csv (str): Ścieżka do pliku CSV z etykietami.
            augment (bool): Czy stosować augmentację danych.
            mean (numpy.ndarray): Średnia pikseli dla normalizacji.
            std (numpy.ndarray): Odchylenie standardowe pikseli dla normalizacji.

        Zwraca:
            None
        """
        self.images_folder = images_folder
        self.labels_df = pd.read_csv(labels_csv)
        self.labels = self.labels_df['grade'].astype(int).values
        self.filenames = self.labels_df['filename'].values
        self.augment = augment
        if augment:
            self.transforms = A.Compose([
                A.HorizontalFlip(p=0.3),
                A.GaussianBlur(blur_limit=(3, 7), p=0.4),
                A.MotionBlur(blur_limit=5, p=0.4),
                A.VerticalFlip(p=0.3),
                A.RandomBrightnessContrast(p=0.4),
                A.Rotate(limit=10, p=0.1),
                A.Affine(scale=(0.8, 1.2), translate_percent=(0.0, 0.0), rotate=(-45, 45), shear=(-20, 20), p=0.5),
                A.Normalize(mean=mean.tolist(), std=std.tolist()),
                ToTensorV2()
            ])
        else:
            self.transforms = A.Compose([
                A.Resize(224, 224),
                A.Normalize(mean=mean.tolist(), std=std.tolist()),
                ToTensorV2()
            ])

    def __len__(self):
        """
        Zwraca długość datasetu.

        Zwraca:
            int: Liczba próbek w zbiorze.
        """
        return len(self.labels)

    def __getitem__(self, idx):
        """
        Zwraca pojedynczy element (obraz i etykietę) z datasetu.

        Parametry:
            idx (int): Indeks próbki.

        Zwraca:
            tuple: (image, label), gdzie image to obraz po przekształceniach, a label to etykieta.
        """
        filename = self.filenames[idx]
        image_path = os.path.join(self.images_folder, filename)
        image = cv2.imread(image_path, cv2.IMREAD_GRAYSCALE)
        if image is None:
            raise FileNotFoundError(f"Unable to read image: {image_path}")
        image = np.dstack([image] * 3).astype(np.float32) / 255.0
        transformed = self.transforms(image=image)
        image = transformed["image"]
        label = torch.tensor(self.labels[idx], dtype=torch.long)
        return image, label


class VisionTransformerModel(nn.Module):
    """
    Klasa modelu Vision Transformer (ViT) z warstwą klasyfikującą.
    """

    def __init__(self, num_classes):
        """
        Inicjalizuje model Vision Transformer i dodaje własną warstwę klasyfikującą.

        Parametry:
            num_classes (int): Liczba klas do przewidywania.

        Zwraca:
            None
        """
        super(VisionTransformerModel, self).__init__()
        self.vit = ViTModel.from_pretrained("google/vit-base-patch16-224")
        vit_hidden_size = self.vit.config.hidden_size
        self.classifier = nn.Sequential(
            nn.Linear(vit_hidden_size, 256),
            nn.LeakyReLU(0.1),
            nn.Dropout(0.4),
            nn.Linear(256, 128),
            nn.LeakyReLU(0.1),
            nn.Dropout(0.3),
            nn.Linear(128, num_classes)
        )

    def forward(self, images):
        """
        Definiuje przepływ danych przez sieć (forward pass).

        Parametry:
            images (torch.Tensor): Tensor wejściowy reprezentujący obrazy.

        Zwraca:
            torch.Tensor: Logity (wynik klasyfikacji) dla każdej próbki.
        """
        outputs = self.vit(pixel_values=images)
        features = outputs.last_hidden_state[:, 0, :]
        logits = self.classifier(features)
        return logits


class ExcelLogger:
    """
    Klasa służąca do logowania wyników treningu i walidacji w pliku Excel.
    """

    def __init__(self, output_path):
        """
        Inicjalizuje obiekt ExcelLogger i tworzy/zaktualizuje arkusze w pliku Excel.

        Parametry:
            output_path (str): Ścieżka do pliku Excel z logami.

        Zwraca:
            None
        """
        self.output_path = output_path
        self.sheet_names = [
            "Class Distribution",
            "Augmentation Info",
            "Training and Validation",
            "Test Results",
            "Per-Class Accuracy",
            "Processed_Data"
        ]
        if os.path.exists(output_path):
            self.workbook = load_workbook(output_path)
            for sheet in self.sheet_names:
                if sheet not in self.workbook.sheetnames:
                    print(f"Tworzenie brakującego arkusza: {sheet}")
                    self.workbook.create_sheet(sheet)
        else:
            self.workbook = Workbook()
            default_sheet = self.workbook.active
            self.workbook.remove(default_sheet)
            for sheet in self.sheet_names:
                self.workbook.create_sheet(sheet)
            self.save()

    def save_class_distribution(self, train_counts, val_counts, test_counts):
        """
        Zapisuje informacje o rozkładzie klas w zbiorach treningowym, walidacyjnym i testowym.

        Parametry:
            train_counts (dict): Rozkład klas w zbiorze treningowym.
            val_counts (dict): Rozkład klas w zbiorze walidacyjnym.
            test_counts (dict): Rozkład klas w zbiorze testowym.

        Zwraca:
            None
        """
        df = pd.DataFrame({
            "Class": ["Grade 1", "Grade 2", "Grade 3"],
            "Train Count (Before Augmentation)": [
                train_counts.get(1, 0),
                train_counts.get(2, 0),
                train_counts.get(3, 0)
            ],
            "Validation Count": [
                val_counts.get(0, 0),
                val_counts.get(1, 0),
                val_counts.get(2, 0)
            ],
            "Test Count": [
                test_counts.get(0, 0),
                test_counts.get(1, 0),
                test_counts.get(2, 0)
            ],
        })
        self._write_to_sheet("Class Distribution", df)

    def save_augmentation_info(self, augmentation_info):
        """
        Zapisuje informacje o augmentacji danych w pliku Excel.

        Parametry:
            augmentation_info (dict): Słownik zawierający szczegóły augmentacji.

        Zwraca:
            None
        """
        df = pd.DataFrame(augmentation_info)
        self._write_to_sheet("Augmentation Info", df)

    def save_epoch_metrics(self, epoch, train_loss, train_acc, val_loss, val_acc):
        """
        Zapisuje metryki (strata i dokładność) dla każdej epoki treningu i walidacji.

        Parametry:
            epoch (int): Numer epoki.
            train_loss (float): Strata treningowa.
            train_acc (float): Dokładność treningowa.
            val_loss (float): Strata walidacyjna.
            val_acc (float): Dokładność walidacyjna.

        Zwraca:
            None
        """
        sheet_name = "Training and Validation"
        sheet = self.workbook[sheet_name]
        if sheet.max_row == 1 and sheet.max_column == 1 and sheet.cell(row=1, column=1).value is None:
            headers = ["Epoch", "Train Loss", "Train Accuracy", "Validation Loss", "Validation Accuracy"]
            sheet.append(headers)
        sheet.append([epoch, train_loss, train_acc, val_loss, val_acc])

    def save_test_results(self, test_loss, test_acc):
        """
        Zapisuje metryki (strata i dokładność) dla zbioru testowego.

        Parametry:
            test_loss (float): Strata testowa.
            test_acc (float): Dokładność testowa.

        Zwraca:
            None
        """
        sheet_name = "Test Results"
        sheet = self.workbook[sheet_name]
        if sheet.max_row == 1 and sheet.max_column == 1 and sheet.cell(row=1, column=1).value is None:
            headers = ["Test Loss", "Test Accuracy"]
            sheet.append(headers)
        sheet.append([test_loss, test_acc])

    def save_class_accuracies(self, epoch, class_accuracies, dataset_type="Validation"):
        """
        Zapisuje dokładności per-klasa w osobnym arkuszu dla danego epoki i typu zbioru (walidacja/test).

        Parametry:
            epoch (int): Numer epoki.
            class_accuracies (dict): Dokładność dla każdej klasy.
            dataset_type (str): Typ zbioru, np. "Validation" lub "Test".

        Zwraca:
            None
        """
        sheet_name = "Per-Class Accuracy"
        sheet = self.workbook[sheet_name]
        if sheet.max_row == 1 and sheet.max_column == 1 and sheet.cell(row=1, column=1).value is None:
            headers = ["Epoch", "Dataset_Type", "Class", "Accuracy (%)"]
            sheet.append(headers)
        for cls, acc in class_accuracies.items():
            sheet.append([epoch, dataset_type, f"Grade {cls + 1}", acc])

    def save(self):
        """
        Zapisuje zmiany w pliku Excel.

        Zwraca:
            None
        """
        self.workbook.save(self.output_path)

    def _write_to_sheet(self, sheet_name, dataframe, append=False):
        """
        Pomocnicza metoda do zapisywania ramki danych (DataFrame) w określonym arkuszu Excel.

        Parametry:
            sheet_name (str): Nazwa arkusza.
            dataframe (pandas.DataFrame): Dane do zapisania.
            append (bool): Czy dopisać do istniejących danych, czy nadpisać.

        Zwraca:
            None
        """
        sheet = self.workbook[sheet_name]
        if not append:
            if sheet.max_row == 1 and sheet.max_column == 1 and sheet.cell(row=1, column=1).value is None:
                for row in dataframe_to_rows(dataframe, index=False, header=True):
                    sheet.append(row)
            else:
                for row in dataframe_to_rows(dataframe, index=False, header=False):
                    sheet.append(row)
        else:
            for row in dataframe_to_rows(dataframe, index=False, header=False):
                sheet.append(row)


class FocalLoss(nn.Module):
    """
    Klasa implementująca funkcję straty Focal Loss, używaną do radzenia sobie z problemem niezrównoważonych klas.
    """

    def __init__(self, alpha=None, gamma=3, reduction='mean'):
        """
        Inicjalizuje Focal Loss.

        Parametry:
            alpha (torch.Tensor lub float lub None): Współczynnik ważenia klas.
            gamma (float): Parametr ogniskowania.
            reduction (str): Redukcja ("mean", "sum" lub "none").

        Zwraca:
            None
        """
        super(FocalLoss, self).__init__()
        self.alpha = alpha if alpha is not None else torch.tensor(1.0)
        self.gamma = gamma
        self.reduction = reduction

    def forward(self, inputs, targets):
        """
        Oblicza stratę Focal Loss.

        Parametry:
            inputs (torch.Tensor): Wyjście modelu (logity).
            targets (torch.Tensor): Prawdziwe etykiety.

        Zwraca:
            torch.Tensor: Wartość straty.
        """
        BCE_loss = nn.CrossEntropyLoss(reduction='none')(inputs, targets)
        pt = torch.exp(-BCE_loss)
        if isinstance(self.alpha, torch.Tensor):
            alpha_t = self.alpha[targets]
        else:
            alpha_t = self.alpha
        F_loss = alpha_t * (1 - pt) ** self.gamma * BCE_loss
        if self.reduction == 'mean':
            return F_loss.mean()
        elif self.reduction == 'sum':
            return F_loss.sum()
        else:
            return F_loss


def validate_network(model, val_loader, device, criterion):
    """
    Waliduje model na zbiorze walidacyjnym/testowym.

    Parametry:
        model (nn.Module): Model PyTorch do walidacji.
        val_loader (DataLoader): DataLoader dla zbioru walidacyjnego/testowego.
        device (torch.device): Urządzenie (CPU/GPU) do obliczeń.
        criterion (nn.Module): Funkcja straty.

    Zwraca:
        tuple: (val_loss, val_acc, class_accuracies)
               val_loss (float) - średnia strata
               val_acc (float) - dokładność
               class_accuracies (dict) - dokładność per-klasa
    """
    model.eval()
    val_loss = 0.0
    correct = 0
    total = 0
    all_preds = []
    all_labels = []
    with torch.no_grad():
        for (x_val, y_val) in val_loader:
            x_val, y_val = x_val.to(device), y_val.to(device)
            outputs = model(x_val)
            loss = criterion(outputs, y_val)
            val_loss += loss.item() * y_val.size(0)
            _, predicted = outputs.max(1)
            correct += predicted.eq(y_val).sum().item()
            total += y_val.size(0)
            all_preds.append(predicted.cpu().numpy())
            all_labels.append(y_val.cpu().numpy())
    val_loss /= total
    val_acc = correct / total
    all_preds = np.concatenate(all_preds)
    all_labels = np.concatenate(all_labels)
    report = classification_report(all_labels, all_preds, target_names=["Grade 1", "Grade 2", "Grade 3"])
    print(report)
    classes = np.unique(all_labels)
    class_accuracies = {}
    for cls in classes:
        cls_mask = (all_labels == cls)
        cls_acc = (all_preds[cls_mask] == cls).mean() if np.any(cls_mask) else 0.0
        class_accuracies[cls] = cls_acc * 100.0
    return val_loss, val_acc, class_accuracies


class EarlyStopping:
    """
    Klasa implementująca mechanizm wczesnego zatrzymywania trenowania (early stopping).
    """

    def __init__(self, patience=5, verbose=False, delta=0, path='best_modelSPRAW.pt'):
        """
        Inicjalizuje obiekt EarlyStopping.

        Parametry:
            patience (int): Liczba epok do wstrzymania, jeśli nie ma poprawy.
            verbose (bool): Czy wypisywać komunikaty o postępie.
            delta (float): Minimalna zmiana w metryce, uznawana za poprawę.
            path (str): Ścieżka do zapisu najlepszego modelu.

        Zwraca:
            None
        """
        self.patience = patience
        self.verbose = verbose
        self.counter = 0
        self.best_score = None
        self.early_stop = False
        self.delta = delta
        self.path = path
        self.best_val_loss = np.Inf

    def __call__(self, val_loss, model):
        """
        Sprawdza, czy wystąpiła poprawa metryki i ewentualnie zapisuje model.
        Jeśli nie ma poprawy przez określoną liczbę epok, trening jest przerywany.

        Parametry:
            val_loss (float): Wartość straty walidacyjnej w danej epoce.
            model (nn.Module): Model PyTorch.

        Zwraca:
            None
        """
        score = -val_loss
        if self.best_score is None:
            self.best_score = score
            self.save_checkpoint(val_loss, model)
        elif score < self.best_score + self.delta:
            self.counter += 1
            if self.verbose:
                print(f'EarlyStopping counter: {self.counter} out of {self.patience}')
            if self.counter >= self.patience:
                self.early_stop = True
        else:
            self.best_score = score
            self.save_checkpoint(val_loss, model)
            self.counter = 0

    def save_checkpoint(self, val_loss, model):
        """
        Zapisuje model, gdy osiągnięto lepszy wynik walidacji.

        Parametry:
            val_loss (float): Wartość straty walidacyjnej.
            model (nn.Module): Model PyTorch.

        Zwraca:
            None
        """
        if self.verbose:
            print(f'Validation loss decreased ({self.best_val_loss:.6f} --> {val_loss:.6f}).  Saving model ...')
        torch.save(model.state_dict(), self.path)
        self.best_val_loss = val_loss


def train_network(model, optimizer, criterion, train_loader, val_loader, device, scheduler, logger,
                  num_epochs=21, early_stopping_patience=5, unfreeze_epoch=10):
    """
    Trenuje model Vision Transformer z wykorzystaniem podanego optymalizatora i scheduler'a.
    Zapisuje metryki treningu i walidacji do loggera Excel. Wprowadza mechanizm early stopping.

    Parametry:
        model (nn.Module): Model do trenowania.
        optimizer (torch.optim.Optimizer): Optymalizator.
        criterion (nn.Module): Funkcja straty.
        train_loader (DataLoader): DataLoader dla zbioru treningowego.
        val_loader (DataLoader): DataLoader dla zbioru walidacyjnego.
        device (torch.device): Urządzenie obliczeniowe (CPU/GPU).
        scheduler (torch.optim.lr_scheduler): Scheduler do zmiany tempa uczenia.
        logger (ExcelLogger): Logger do zapisywania wyników w Excel.
        num_epochs (int): Maksymalna liczba epok treningu.
        early_stopping_patience (int): Liczba epok braku poprawy, po której kończymy trening.
        unfreeze_epoch (int): Numer epoki, po którym odblokowujemy warstwy ViT.

    Zwraca:
        nn.Module: Najlepszy wytrenowany model (wczytany ze stanu zapisów).
    """
    early_stopping = EarlyStopping(patience=early_stopping_patience, verbose=True, path='best_modelSPRAW.pt')
    for epoch in range(num_epochs):
        if early_stopping.early_stop:
            print("Early stopping triggered.")
            break
        model.train()
        train_loss = 0.0
        correct = 0
        total = 0
        with tqdm(train_loader, desc=f"Epoch {epoch + 1}/{num_epochs} [Training]") as pbar:
            for images_batch, labels_batch in pbar:
                images_batch, labels_batch = images_batch.to(device), labels_batch.to(device)
                optimizer.zero_grad()
                outputs = model(images_batch)
                loss = criterion(outputs, labels_batch)
                loss.backward()
                optimizer.step()
                train_loss += loss.item() * labels_batch.size(0)
                _, predicted = outputs.max(1)
                correct += predicted.eq(labels_batch).sum().item()
                total += labels_batch.size(0)
                pbar.set_postfix(loss=f"{train_loss / total:.4f}", accuracy=f"{correct / total:.4f}")
        train_acc = correct / total
        train_loss /= total
        print(f'Train Loss: {train_loss:.3f} | Train Accuracy: {train_acc:.3f}')
        val_loss, val_acc, val_class_accuracies = validate_network(model, val_loader, device, criterion)
        logger.save_epoch_metrics(epoch + 1, train_loss, train_acc, val_loss, val_acc)
        logger.save_class_accuracies(epoch + 1, val_class_accuracies, dataset_type="Validation")
        print(f"Validation: Loss: {val_loss:.4f}, Acc: {val_acc:.4f}")
        print("Dokładność per-klasa (walidacja):")
        for cls_id, acc_cls in val_class_accuracies.items():
            print(f"Grade {cls_id + 1}: {acc_cls:.2f}%")
        if isinstance(scheduler, torch.optim.lr_scheduler.ReduceLROnPlateau):
            scheduler.step(val_loss)
        else:
            scheduler.step()
        early_stopping(val_loss, model)
        if early_stopping.early_stop:
            print("Early stopping")
            break
        if epoch + 1 == unfreeze_epoch:
            print("Odblokowywanie warstw ViT do trenowania.")
            for param in model.vit.parameters():
                param.requires_grad = True
            optimizer = torch.optim.Adam(model.parameters(), lr=5e-8, weight_decay=1e-4)
    model.load_state_dict(torch.load('best_modelSPRAW.pt'))
    return model


def visualize_sample_images(original_image, augmented_image):
    """
    Wizualizuje oryginalny i zaugmentowany obraz obok siebie.

    Parametry:
        original_image (numpy.ndarray): Oryginalny obraz w formacie [W, H].
        augmented_image (numpy.ndarray): Zaugmentowany obraz w formacie [W, H].

    Zwraca:
        None
    """
    plt.figure(figsize=(6, 3))
    plt.subplot(1, 2, 1)
    plt.imshow(original_image, cmap='gray')
    plt.title('Oryginalny obraz')
    plt.axis('off')
    plt.subplot(1, 2, 2)
    plt.imshow(augmented_image, cmap='gray')
    plt.title('Obraz po augmentacji')
    plt.axis('off')
    plt.tight_layout()
    plt.show()


def main():
    """
    Główna funkcja skryptu:
    1. Tworzy obiekt MRIDataProcessor i (w razie potrzeby) przetwarza dane.
    2. Ładuje obrazy i etykiety, dzieli je na zbiory (train, val, test).
    3. Zapisuje wyniki w pliku Excel i inicjalizuje loggera.
    4. Przeprowadza augmentację danych, aby zbalansować klasy w zbiorze treningowym.
    5. Tworzy i scala dataset oryginalny i zaugmentowany, oblicza wagi klas.
    6. Tworzy i trenuje model Vision Transformer z wykorzystaniem Focal Loss i early stopping.
    7. Ocena modelu na zbiorze testowym i zapisanie wyników do pliku Excel.

    Zwraca:
        None
    """
    processed_folder = r"C:\Users\a\Desktop\MRI\processed_images"
    train_folder = os.path.join(processed_folder, "train")
    val_folder = os.path.join(processed_folder, "validation")
    test_folder = os.path.join(processed_folder, "test")
    output_excel_path = r"C:\Users\a\Desktop\MRI\Model26_TYLKOSPRAWDZAM.xlsx"
    os.makedirs(processed_folder, exist_ok=True)
    processor = MRIDataProcessor(
        base_path=r"C:\Users\a\Desktop\MRI\manifest-1734817926430\Duke-Breast-Cancer-MRI",
        roi_excel_path=r"C:\Users\a\Desktop\MRI\Annotation_Boxes.xlsx",
        clinical_excel_path=r"C:\Users\a\Desktop\MRI\Clinical_and_Other_Features.xlsx",
        processed_folder=processed_folder,
        output_excel_path=output_excel_path,
    )
    labels_csv_path = os.path.join(processed_folder, "labels.csv")
    if not os.path.exists(labels_csv_path):
        print("Labels file not found. Starting data processing...")
        processor.process_and_save_data()
    else:
        print("Labels file found. Proceeding to next steps.")
    images, labels, class_counts = processor.load_filtered_data()
    label_mapping = {1: 0, 2: 1, 3: 2}
    assert labels.max() < 3, f"Found label {labels.max()} which is out of bounds."
    assert labels.min() >= 0, f"Found label {labels.min()} which is out of bounds."
    print("Unique labels in the dataset after mapping:", np.unique(labels))
    x_train, x_val, x_test, y_train, y_val, y_test = processor.ensure_train_test_split(
        images, labels, test_size=0.1, val_size=0.1
    )
    print("\nClass distribution in Validation Set:")
    unique_val, counts_val = np.unique(y_val, return_counts=True)
    for cls, count in zip(unique_val, counts_val):
        print(f"Grade {cls + 1}: {count} samples")
    print("\nClass distribution in Test Set:")
    unique_test, counts_test = np.unique(y_test, return_counts=True)
    for cls, count in zip(unique_test, counts_test):
        print(f"Grade {cls + 1}: {count} samples")
    logger = ExcelLogger(output_excel_path)
    logger.save_class_distribution(
        train_counts=class_counts,
        val_counts=dict(zip(unique_val, counts_val)),
        test_counts=dict(zip(unique_test, counts_test)),
    )
    print("\nSaving validation and test data to separate folders...")
    processor.save_split_data(x_val, y_val, val_folder, os.path.join(val_folder, "labels_val.csv"))
    processor.save_split_data(x_test, y_test, test_folder, os.path.join(test_folder, "labels_test.csv"))
    print("\nStarting training set augmentation to balance classes...")
    os.makedirs(train_folder, exist_ok=True)
    max_count = max([class_counts.get(cls, 0) for cls in [1, 2, 3]])
    print(f"Maximum number of samples in a class: {max_count}")
    augmentation_info = {"Set": [], "Class": [], "Before_Augmentation": [], "After_Augmentation": []}
    for original_class, mapped_class in label_mapping.items():
        current_count = class_counts.get(original_class, 0)
        required_aug = max_count - current_count
        augmentation_info["Set"].append("Train")
        augmentation_info["Class"].append(f"Grade {mapped_class + 1}")
        augmentation_info["Before_Augmentation"].append(current_count)
        if required_aug > 0:
            print(f"Augmenting Grade {mapped_class + 1} with {required_aug} new samples.")
            class_images = x_train[y_train == mapped_class]
            if len(class_images) == 0:
                print(f"No images found for Grade {mapped_class + 1}. Skipping augmentation.")
                augmentation_info["After_Augmentation"].append(0)
                continue
            transform = A.Compose([
                A.HorizontalFlip(p=0.5),
                A.VerticalFlip(p=0.5),
                A.RandomBrightnessContrast(p=0.3),
                A.Rotate(limit=30, p=0.3),
            ])
            augmented_images = []
            augmented_labels = []
            for i in range(required_aug):
                img = class_images[np.random.randint(0, len(class_images))]
                augmented = transform(image=img)
                augmented_image = augmented["image"]
                augmented_images.append(augmented_image)
                filename = f"aug_grade_{mapped_class + 1}_{i}.png"
                save_path = os.path.join(train_folder, filename)
                augmented_image_uint8 = (augmented_image * 255).astype(np.uint8)
                cv2.imwrite(save_path, augmented_image_uint8)
                augmented_labels.append(mapped_class)
                augmented_labels_df = pd.DataFrame({
                    'filename': [filename],
                    'grade': [mapped_class]
                })
                csv_path = os.path.join(train_folder, "labels_augmented.csv")
                if not os.path.exists(csv_path):
                    augmented_labels_df.to_csv(csv_path, index=False, header=True)
                else:
                    augmented_labels_df.to_csv(csv_path, mode='a', index=False, header=False)
                if i < 5:
                    original_image = img * 255.0
                    augmented_image_display = augmented_image
                    visualize_sample_images(original_image.astype(np.uint8), augmented_image_display.astype(np.uint8))
            augmentation_info["After_Augmentation"].append(required_aug)
        else:
            print(f"No augmentation needed for Grade {mapped_class + 1}.")
            augmentation_info["After_Augmentation"].append(0)
    logger.save_augmentation_info(augmentation_info)
    print("\nClass distribution in Training Set before augmentation:")
    for original_class, mapped_class in label_mapping.items():
        count = class_counts.get(original_class, 0)
        print(f"Grade {mapped_class + 1}: {count} samples")
    print("\nClass distribution in Training Set after augmentation:")
    augmented_labels_path = os.path.join(train_folder, "labels_augmented.csv")
    if os.path.exists(augmented_labels_path):
        augmented_labels_df = pd.read_csv(augmented_labels_path)
        augmented_counts = augmented_labels_df['grade'].value_counts().to_dict()
        for mapped_class in range(3):
            total = class_counts.get(mapped_class + 1, 0) + augmented_counts.get(mapped_class, 0)
            print(f"Grade {mapped_class + 1}: {total} samples")
    else:
        print("No augmented labels found.")
    print("\nLoading original training data...")
    original_labels_csv_path = os.path.join(train_folder, "labels_train.csv")
    if not os.path.exists(original_labels_csv_path):
        print("Saving original training data to labels_train.csv...")
        processor.save_split_data(x_train, y_train, train_folder, original_labels_csv_path)
    else:
        print("Original training labels already exist.")
    original_train_dataset = BreastCancerDataset(
        images_folder=train_folder,
        labels_csv=original_labels_csv_path,
        augment=False,
        mean=processor.calculate_mean_std(x_train)[0],
        std=processor.calculate_mean_std(x_train)[1]
    )
    if os.path.exists(augmented_labels_path):
        augmented_train_dataset = BreastCancerDataset(
            images_folder=train_folder,
            labels_csv=augmented_labels_path,
            augment=False,
            mean=processor.calculate_mean_std(x_train)[0],
            std=processor.calculate_mean_std(x_train)[1]
        )
    else:
        print("No augmented data found. Using original training data only.")
        augmented_train_dataset = None
    if augmented_train_dataset:
        combined_train_dataset = ConcatDataset([original_train_dataset, augmented_train_dataset])
    else:
        combined_train_dataset = original_train_dataset
    all_train_labels = []
    for _, label in combined_train_dataset:
        all_train_labels.append(label.item())
    all_train_labels = np.array(all_train_labels)
    unique_train_labels = np.unique(all_train_labels)
    print(f"\nUnikalne etykiety w zbiorze treningowym: {unique_train_labels}")
    if len(unique_train_labels) < 3:
        print("WARNING: Not all classes are present in the training set!")
    else:
        print("All classes are present in the training set.")
    class_weights = compute_class_weight("balanced", classes=np.unique(all_train_labels), y=all_train_labels)
    class_weights = torch.FloatTensor(class_weights).to(torch.device("cuda" if torch.cuda.is_available() else "cpu"))
    sample_weights = class_weights[all_train_labels]
    sampler = WeightedRandomSampler(weights=sample_weights, num_samples=len(sample_weights), replacement=True)
    criterion = FocalLoss(alpha=class_weights, gamma=3, reduction="mean")
    device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
    print(f"\nUsing device: {device}")
    model = VisionTransformerModel(num_classes=3).to(device)
    optimizer = torch.optim.Adam(model.parameters(), lr=1e-7, weight_decay=1e-4)
    scheduler = torch.optim.lr_scheduler.ReduceLROnPlateau(optimizer, mode='min', factor=0.1, patience=3, verbose=True)
    for param in model.vit.parameters():
        param.requires_grad = False
    print("\nCreating Datasets and Dataloaders...")
    train_loader = DataLoader(combined_train_dataset, batch_size=8, sampler=sampler)
    val_dataset = BreastCancerDataset(
        images_folder=val_folder,
        labels_csv=os.path.join(val_folder, "labels_val.csv"),
        augment=False,
        mean=processor.calculate_mean_std(x_train)[0],
        std=processor.calculate_mean_std(x_train)[1]
    )
    test_dataset = BreastCancerDataset(
        images_folder=test_folder,
        labels_csv=os.path.join(test_folder, "labels_test.csv"),
        augment=False,
        mean=processor.calculate_mean_std(x_train)[0],
        std=processor.calculate_mean_std(x_train)[1]
    )
    val_loader = DataLoader(val_dataset, batch_size=8, shuffle=False)
    test_loader = DataLoader(test_dataset, batch_size=8, shuffle=False)

    def check_classes_present(labels, split_name):
        """
        Sprawdza, czy wszystkie klasy są obecne w danym podzbiorze.

        Parametry:
            labels (numpy.ndarray): Etykiety w zbiorze.
            split_name (str): Nazwa podzbioru (np. "Treningowy", "Walidacyjny", "Testowy").

        Zwraca:
            None
        """
        unique = np.unique(labels)
        print(f"\nUnikalne etykiety w zbiorze {split_name}: {unique}")
        if len(unique) < 3:
            print(f"WARNING: Nie wszystkie klasy są obecne w zbiorze {split_name}!")

    check_classes_present(y_train, "Treningowy")
    check_classes_present(y_val, "Walidacyjny")
    check_classes_present(y_test, "Testowy")
    print("\nStarting model training...")
    trained_model = train_network(
        model, optimizer, criterion, train_loader, val_loader,
        device, scheduler, logger,
        num_epochs=100,
        early_stopping_patience=5,
        unfreeze_epoch=10
    )
    print("\nEvaluating on the test set:")
    test_loss, test_acc, test_class_accuracies = validate_network(trained_model, test_loader, device, criterion)
    logger.save_test_results(test_loss, test_acc)
    logger.save()
    print("\nTraining and testing completed. All logs saved to Excel.")
    print("\nFinal Class Distributions:")
    print("Validation Set:")
    print(f"Grade 1: {dict(zip(unique_val, counts_val)).get(0, 0)} samples")
    print(f"Grade 2: {dict(zip(unique_val, counts_val)).get(1, 0)} samples")
    print(f"Grade 3: {dict(zip(unique_val, counts_val)).get(2, 0)} samples")
    print("\nTest Set:")
    print(f"Grade 1: {dict(zip(unique_test, counts_test)).get(0, 0)} samples")
    print(f"Grade 2: {dict(zip(unique_test, counts_test)).get(1, 0)} samples")
    print(f"Grade 3: {dict(zip(unique_test, counts_test)).get(2, 0)} samples")


if __name__ == "__main__":
    main()
