import os
from collections import defaultdict

import pandas as pd
import numpy as np
import cv2
import torch
import torch.nn as nn
from torch.utils.data import Dataset, DataLoader, WeightedRandomSampler, ConcatDataset
from transformers import ViTModel
from sklearn.model_selection import train_test_split
from sklearn.metrics import classification_report
from sklearn.utils.class_weight import compute_class_weight
from tqdm import tqdm
import albumentations as A
from albumentations.pytorch import ToTensorV2
import matplotlib.pyplot as plt
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import pydicom


class MRIDataProcessor:
    """
    Klasa do przetwarzania danych MRI pod kątem wyznaczania i zapisu obszarów zainteresowania (ROI)
    oraz przygotowania etykiet (molecular subtypes) do dalszej analizy.

    Atrybuty:
        base_path (str): Ścieżka do głównego folderu z plikami DICOM.
        roi_excel_path (str): Ścieżka do pliku Excel zawierającego informacje o ROI.
        clinical_excel_path (str): Ścieżka do pliku Excel zawierającego dane kliniczne.
        processed_folder (str): Ścieżka do folderu, w którym zostaną zapisane przetworzone obrazy.
        output_excel_path (str): Ścieżka do pliku Excel z podsumowaniem przetwarzania.
    """

    def __init__(self, base_path, roi_excel_path, clinical_excel_path, processed_folder, output_excel_path):
        """
        Inicjalizuje obiekt klasy MRIDataProcessor.

        Parametry:
            base_path (str): Ścieżka do głównego folderu z plikami DICOM.
            roi_excel_path (str): Ścieżka do pliku Excel z ROI.
            clinical_excel_path (str): Ścieżka do pliku Excel z danymi klinicznymi.
            processed_folder (str): Ścieżka do folderu na przetworzone obrazy.
            output_excel_path (str): Ścieżka do pliku Excel z podsumowaniem przetwarzania.
        """
        self.base_path = base_path
        self.roi_excel_path = roi_excel_path
        self.clinical_excel_path = clinical_excel_path
        self.processed_folder = processed_folder
        self.output_excel_path = output_excel_path
        os.makedirs(self.processed_folder, exist_ok=True)

    def load_roi_and_slices_from_excel(self, patient_id):
        """
        Wczytuje z pliku Excel współrzędne (ROI) oraz zakres warstw (slice range) dla danego pacjenta.

        Parametry:
            patient_id (str): Identyfikator pacjenta.

        Zwraca:
            tuple: (start_row, end_row, start_column, end_column, start_slice, end_slice)

        Wyjątki:
            ValueError: Gdy nie znaleziono danych ROI dla danego pacjenta.
        """
        df = pd.read_excel(self.roi_excel_path)
        roi_data = df[df["Patient ID"] == patient_id]
        if roi_data.empty:
            raise ValueError(f"Brak danych ROI dla pacjenta {patient_id}")
        return (
            roi_data["Start Row"].values[0] - 10,
            roi_data["End Row"].values[0] + 10,
            roi_data["Start Column"].values[0] - 10,
            roi_data["End Column"].values[0] + 10,
            roi_data["Start Slice"].values[0],
            roi_data["End Slice"].values[0],
        )

    def load_clinical_labels(self, patient_id):
        """
        Wczytuje z pliku Excel dane kliniczne (tumor grade i molecular subtype) dla danego pacjenta.

        Parametry:
            patient_id (str): Identyfikator pacjenta.

        Zwraca:
            tuple: (grade, subtype) lub (None, None) jeśli brak danych.
        """
        df = pd.read_excel(self.clinical_excel_path, header=None)
        df.columns = [' '.join(map(str, col)).replace("\n", " ").strip() for col in zip(*df.iloc[:3].values)]
        df = df[3:]
        patient_id_col = [col for col in df.columns if "Breast_MRI" in str(df[col].values)]
        tumor_grade_col = [col for col in df.columns if "Tumor Grade(N)" in col]
        mol_subtype_col = [col for col in df.columns if "Mol Subtype" in col]
        clinical_data = df[df[patient_id_col[0]] == patient_id]
        if clinical_data.empty:
            return None, None
        return clinical_data[tumor_grade_col[0]].values[0], clinical_data[mol_subtype_col[0]].values[0]

    def process_and_save_data(self):
        """
        Przetwarza dane DICOM, wycina obszar ROI i dokonuje wstępnych transformacji (medianBlur, CLAHE, normalizacja).
        Zapisuje przetworzone obrazy do pliku PNG oraz tworzy plik CSV (labels.csv) z etykietami.
        Dodatkowo zapisuje podsumowanie do pliku Excel.

        Zwraca:
            None
        """
        data_summary = []
        labels = []
        for patient_id in tqdm(os.listdir(self.base_path), desc="Przetwarzanie pacjentów"):
            try:
                image_filename = f"{patient_id}.png"
                processed_path = os.path.join(self.processed_folder, image_filename)
                if os.path.exists(processed_path):
                    print(f"Dane dla pacjenta {patient_id} już istnieją. Pomijanie...")
                    continue

                start_row, end_row, start_column, end_column, start_slice, end_slice = self.load_roi_and_slices_from_excel(
                    patient_id
                )
                dicom_folder = self.find_matching_dicom_folder(patient_id)
                dicom_files = sorted(
                    [os.path.join(dicom_folder, f) for f in os.listdir(dicom_folder) if f.endswith(".dcm")]
                )
                slice_range = dicom_files[start_slice:end_slice + 1]
                if not slice_range:
                    print(f"Brak plików DICOM w zakresie dla pacjenta {patient_id}. Pomijanie...")
                    continue

                middle_index = len(slice_range) // 2
                dicom_data = pydicom.dcmread(slice_range[middle_index])
                image = dicom_data.pixel_array

                roi_image = image[start_row:end_row, start_column:end_column]
                roi_image = cv2.medianBlur(roi_image, 3)
                clahe = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(8, 8))
                normalized_image = cv2.normalize(roi_image, None, 0, 255, cv2.NORM_MINMAX).astype(np.uint8)
                processed_image = clahe.apply(normalized_image)
                processed_image = cv2.resize(processed_image, (224, 224))
                cv2.imwrite(processed_path, processed_image)

                grade, subtype = self.load_clinical_labels(patient_id)
                if grade is not None and subtype is not None:
                    labels.append({'filename': image_filename, 'subtype': int(subtype)})
                    data_summary.append(
                        {"Patient ID": patient_id, "Mol Subtype": int(subtype), "Tumor Grade (Nuclear)": grade}
                    )
            except Exception as e:
                print(f"Nie udało się przetworzyć danych pacjenta {patient_id}: {e}")

        labels_df = pd.DataFrame(labels)
        labels_csv_path = os.path.join(self.processed_folder, "labels.csv")
        labels_df.to_csv(labels_csv_path, index=False)

        df_summary = pd.DataFrame(data_summary)
        with pd.ExcelWriter(self.output_excel_path, engine="openpyxl") as writer:
            df_summary.to_excel(writer, sheet_name="Processed_Data", index=False)
        print("Podsumowanie zapisano do pliku Excel.")

    def find_matching_dicom_folder(self, patient_id):
        """
        Wyszukuje folder zawierający pliki DICOM dla konkretnego pacjenta.

        Parametry:
            patient_id (str): Identyfikator pacjenta.

        Zwraca:
            str: Ścieżka do folderu zawierającego pliki DICOM.

        Wyjątki:
            FileNotFoundError: Gdy folder pacjenta nie istnieje albo nie znaleziono plików DICOM.
        """
        patient_path = os.path.join(self.base_path, patient_id)
        if not os.path.exists(patient_path):
            raise FileNotFoundError(f"Folder pacjenta {patient_id} nie istnieje: {patient_path}")

        dicom_folders = []
        for root, dirs, files in os.walk(patient_path):
            if any(file.endswith('.dcm') for file in files):
                dicom_folders.append(root)

        if not dicom_folders:
            raise FileNotFoundError(f"Nie znaleziono plików DICOM dla pacjenta {patient_id}.")

        return dicom_folders[0]

    def calculate_mean_std(self, images):
        """
        Oblicza średnią (mean) i odchylenie standardowe (std) pikseli dla kolekcji obrazów.

        Parametry:
            images (list or np.ndarray): Lista lub tablica obrazów w formacie [N, W, H, C].

        Zwraca:
            tuple: (mean, std) — Średnie i odchylenie standardowe dla każdego kanału.
        """
        all_pixels = np.concatenate([img.flatten().reshape(-1, 3) for img in images], axis=0)
        mean = np.mean(all_pixels, axis=0) / 255.0
        std = np.std(all_pixels, axis=0) / 255.0
        return mean, std

    def load_filtered_data(self):
        """
        Wczytuje i filtruje przetworzone dane obrazowe oraz etykiety molecular subtype z pliku CSV (labels.csv).

        Zwraca:
            tuple: (images, labels, molecular_subtype_counts)
                   images (list[np.ndarray]): Lista obrazów w formacie float32 [W, H, 3].
                   labels (list[int]): Lista etykiet molecular subtype.
                   molecular_subtype_counts (dict): Zliczenia wystąpień poszczególnych subtypów.
        """
        images, labels = [], []
        molecular_subtype_counts = {}
        labels_csv_path = os.path.join(self.processed_folder, "labels.csv")
        if not os.path.exists(labels_csv_path):
            raise FileNotFoundError(f"Plik z etykietami nie został znaleziony: {labels_csv_path}")

        labels_df = pd.read_csv(labels_csv_path)
        for index, row in tqdm(labels_df.iterrows(), total=labels_df.shape[0],
                               desc="Wczytywanie przetworzonych obrazów"):
            filename = row['filename']
            image_path = os.path.join(self.processed_folder, filename)
            label = row['subtype']
            image = cv2.imread(image_path, cv2.IMREAD_GRAYSCALE)
            if image is None:
                print(f"Nie można wczytać obrazu: {image_path}")
                continue

            image = np.dstack([image] * 3).astype(np.float32) / 255.0
            molecular_subtype = label

            if pd.isna(molecular_subtype):
                print(f"Pomijanie obrazu {filename}, ponieważ Molecular Subtype jest NaN.")
                continue

            try:
                molecular_subtype = int(molecular_subtype)
            except ValueError:
                print(f"Nieprawidłowa wartość molecular subtype dla obrazu {filename}. Pomijanie...")
                continue

            images.append(image)
            labels.append(molecular_subtype)
            molecular_subtype_counts[molecular_subtype] = molecular_subtype_counts.get(molecular_subtype, 0) + 1

        print("\nPodsumowanie liczby przypadków dla każdej klasy molecular_subtype:")
        print(molecular_subtype_counts)
        print(f"Wczytano {len(images)} obrazów i {len(labels)} etykiet.")
        min_pixel = min(img.min() for img in images)
        max_pixel = max(img.max() for img in images)
        print(f"Zakres wartości pikseli przed augmentacją: min={min_pixel}, max={max_pixel}")
        unique_labels = np.unique(labels)
        print(f"Unikalne etykiety po mapowaniu: {unique_labels}")
        return images, labels, molecular_subtype_counts

    def save_split_data(self, images, labels, folder, labels_csv_path):
        """
        Zapisuje obrazy i odpowiadające im etykiety do wskazanego folderu oraz tworzy plik CSV z etykietami.

        Parametry:
            images (list[np.ndarray]): Lista obrazów w formacie [W, H, 3].
            labels (list[int]): Lista etykiet.
            folder (str): Ścieżka do folderu docelowego.
            labels_csv_path (str): Ścieżka do pliku CSV z etykietami.

        Zwraca:
            None
        """
        os.makedirs(folder, exist_ok=True)
        labels_list = []
        for idx, (img, lbl) in enumerate(zip(images, labels)):
            filename = f"{idx}.png"
            img_uint8 = (img * 255).astype(np.uint8)
            image_path = os.path.join(folder, filename)
            cv2.imwrite(image_path, img_uint8)
            labels_list.append({'filename': filename, 'subtype': lbl})

        labels_df = pd.DataFrame(labels_list)
        labels_df.to_csv(labels_csv_path, index=False)
        print(f"Saved {len(images)} images to {folder} with labels at {labels_csv_path}.")

    @staticmethod
    def ensure_train_test_split(images, labels, test_size=0.1, val_size=0.1, random_state=42):
        """
        Dzieli dane na zbiory: treningowy, walidacyjny i testowy z zachowaniem proporcji klas (stratyfikacja).

        Parametry:
            images (list[np.ndarray]): Lista obrazów.
            labels (list[int]): Lista etykiet.
            test_size (float): Ułamek zbioru danych przeznaczony na test.
            val_size (float): Ułamek zbioru danych przeznaczony na walidację.
            random_state (int): Ziarno losowe dla powtarzalności podziału.

        Zwraca:
            tuple: (x_train, x_val, x_test, y_train, y_val, y_test)
        """
        x_train, x_temp, y_train, y_temp = train_test_split(
            images, labels, test_size=test_size + val_size, random_state=random_state, stratify=labels
        )
        val_ratio = val_size / (test_size + val_size)
        x_val, x_test, y_val, y_test = train_test_split(
            x_temp, y_temp, test_size=1 - val_ratio, random_state=random_state, stratify=y_temp
        )

        print(f"Rozmiar zbioru treningowego: {len(x_train)}")
        print(f"Rozmiar zbioru walidacyjnego: {len(x_val)}")
        print(f"Rozmiar zbioru testowego: {len(x_test)}")

        for split_name, split_labels in zip(["Treningowy", "Walidacyjny", "Testowy"], [y_train, y_val, y_test]):
            unique, counts = np.unique(split_labels, return_counts=True)
            print(f"\nRozkład klas w zbiorze {split_name}:")
            for cls, count in zip(unique, counts):
                print(f"Subtype {cls}: {count} próbek")

        return x_train, x_val, x_test, y_train, y_val, y_test


class BreastCancerDataset(Dataset):
    """
    Klasa Dataset dla danych raka piersi.
    Implementuje interfejs Dataset z biblioteki PyTorch.
    """

    def __init__(self, images_folder, labels_csv, augment=False, mean=None, std=None):
        """
        Inicjalizuje dataset, wczytując ścieżki do obrazów i etykiet z pliku CSV.

        Parametry:
            images_folder (str): Ścieżka do folderu z obrazami.
            labels_csv (str): Ścieżka do pliku CSV z etykietami.
            augment (bool): Czy zastosować augmentację danych.
            mean (np.ndarray): Średnie wartości pikseli do normalizacji.
            std (np.ndarray): Odchylenie standardowe pikseli do normalizacji.

        Zwraca:
            None
        """
        self.images_folder = images_folder
        self.labels_df = pd.read_csv(labels_csv)
        self.labels = self.labels_df['subtype'].astype(int).values
        self.filenames = self.labels_df['filename'].values
        self.augment = augment

        if augment:
            self.transforms = A.Compose([
                A.HorizontalFlip(p=0.3),
                A.GaussianBlur(blur_limit=(3, 7), p=0.4),
                A.MotionBlur(blur_limit=5, p=0.4),
                A.VerticalFlip(p=0.3),
                A.RandomBrightnessContrast(p=0.4),
                A.Rotate(limit=10, p=0.1),
                A.Affine(scale=(0.8, 1.2), translate_percent=(0.0, 0.0), rotate=(-45, 45), shear=(-20, 20), p=0.5),
                A.Normalize(mean=mean.tolist(), std=std.tolist()),
                ToTensorV2()
            ])
        else:
            self.transforms = A.Compose([
                A.Resize(224, 224),
                A.Normalize(mean=mean.tolist(), std=std.tolist()),
                ToTensorV2()
            ])

    def __len__(self):
        """
        Zwraca liczbę próbek w zbiorze.

        Zwraca:
            int: Liczba próbek.
        """
        return len(self.labels)

    def __getitem__(self, idx):
        """
        Zwraca pojedynczą próbkę datasetu w postaci (obraz, etykieta).

        Parametry:
            idx (int): Indeks próbki.

        Zwraca:
            tuple: (torch.Tensor, torch.Tensor) — obraz i etykieta.
        """
        filename = self.filenames[idx]
        image_path = os.path.join(self.images_folder, filename)
        image = cv2.imread(image_path, cv2.IMREAD_GRAYSCALE)
        if image is None:
            raise FileNotFoundError(f"Unable to read image: {image_path}")

        image = np.dstack([image] * 3).astype(np.float32) / 255.0
        transformed = self.transforms(image=image)
        image = transformed["image"]
        label = torch.tensor(self.labels[idx], dtype=torch.long)
        return image, label


class VisionTransformerModel(nn.Module):
    """
    Model klasyfikujący oparty na Vision Transformerze (ViT) z dodatkową warstwą klasyfikującą.
    """

    def __init__(self, num_classes):
        """
        Inicjalizuje Vision Transformer z wstępnie wytrenowanymi wagami i dokłada własną warstwę klasyfikacji.

        Parametry:
            num_classes (int): Liczba klas do przewidywania.

        Zwraca:
            None
        """
        super(VisionTransformerModel, self).__init__()
        self.vit = ViTModel.from_pretrained("google/vit-base-patch16-224")
        vit_hidden_size = self.vit.config.hidden_size
        self.classifier = nn.Sequential(
            nn.Linear(vit_hidden_size, 256),
            nn.LeakyReLU(0.1),
            nn.Dropout(0.4),
            nn.Linear(256, 128),
            nn.LeakyReLU(0.1),
            nn.Dropout(0.3),
            nn.Linear(128, num_classes)
        )

    def forward(self, images):
        """
        Definiuje przepływ danych przez sieć.

        Parametry:
            images (torch.Tensor): Tensor o kształcie (batch_size, 3, 224, 224).

        Zwraca:
            torch.Tensor: Logity (przed softmaxem) dla każdej klasy.
        """
        outputs = self.vit(pixel_values=images)
        features = outputs.last_hidden_state[:, 0, :]
        logits = self.classifier(features)
        return logits


class FocalLoss(nn.Module):
    """
    Implementacja funkcji straty Focal Loss, pomocna przy nierównomiernej liczbie próbek w klasach.
    """

    def __init__(self, alpha=None, gamma=3, reduction='mean'):
        """
        Inicjalizuje funkcję straty Focal Loss.

        Parametry:
            alpha (torch.Tensor lub float lub None): Współczynnik skalujący wagi poszczególnych klas.
            gamma (float): Parametr ogniskujący (focusing parameter).
            reduction (str): Typ redukcji ('mean', 'sum', 'none').

        Zwraca:
            None
        """
        super(FocalLoss, self).__init__()
        self.alpha = alpha if alpha is not None else torch.tensor(1.0)
        self.gamma = gamma
        self.reduction = reduction

    def forward(self, inputs, targets):
        """
        Oblicza wartość funkcji straty Focal Loss.

        Parametry:
            inputs (torch.Tensor): Logity wyjściowe modelu.
            targets (torch.Tensor): Prawidłowe etykiety klas.

        Zwraca:
            torch.Tensor: Wartość straty.
        """
        BCE_loss = nn.CrossEntropyLoss(reduction='none')(inputs, targets)
        pt = torch.exp(-BCE_loss)
        if isinstance(self.alpha, torch.Tensor):
            alpha_t = self.alpha[targets]
        else:
            alpha_t = self.alpha
        F_loss = alpha_t * (1 - pt) ** self.gamma * BCE_loss
        if self.reduction == 'mean':
            return F_loss.mean()
        elif self.reduction == 'sum':
            return F_loss.sum()
        else:
            return F_loss


def validate_network(model, val_loader, device, criterion):
    """
    Waliduje model na zbiorze walidacyjnym/testowym, obliczając stratę, dokładność oraz raport
    klasyfikacji (precision, recall) dla każdej klasy.

    Parametry:
        model (nn.Module): Model do walidacji.
        val_loader (DataLoader): Loader zbioru walidacyjnego/testowego.
        device (torch.device): Urządzenie (CPU/GPU).
        criterion (nn.Module): Funkcja straty.

    Zwraca:
        tuple: (val_loss, val_acc, per_class_metrics)
    """
    model.eval()
    val_loss = 0.0
    correct = 0
    total = 0
    all_preds = []
    all_labels = []

    with torch.no_grad():
        for (x_val, y_val) in val_loader:
            x_val, y_val = x_val.to(device), y_val.to(device)
            outputs = model(x_val)
            loss = criterion(outputs, y_val)
            val_loss += loss.item() * y_val.size(0)
            _, predicted = outputs.max(1)
            correct += predicted.eq(y_val).sum().item()
            total += y_val.size(0)
            all_preds.append(predicted.cpu().numpy())
            all_labels.append(y_val.cpu().numpy())

    val_loss /= total
    val_acc = correct / total
    all_preds = np.concatenate(all_preds)
    all_labels = np.concatenate(all_labels)
    report = classification_report(all_labels, all_preds,
                                  target_names=["Subtype 0", "Subtype 1", "Subtype 2", "Subtype 3"],
                                  output_dict=True)
    print(classification_report(all_labels, all_preds,
                                target_names=["Subtype 0", "Subtype 1", "Subtype 2", "Subtype 3"]))
    per_class_metrics = {}
    for cls in ["Subtype 0", "Subtype 1", "Subtype 2", "Subtype 3"]:
        if cls in report:
            per_class_metrics[cls] = {
                'precision': report[cls]['precision'],
                'recall': report[cls]['recall']
            }
        else:
            per_class_metrics[cls] = {
                'precision': 0.0,
                'recall': 0.0
            }
    return val_loss, val_acc, per_class_metrics


class EarlyStopping:
    """
    Mechanizm wczesnego zatrzymania trenowania, jeśli metryka walidacyjna nie poprawia się
    przez określoną liczbę epok.
    """

    def __init__(self, patience=5, verbose=False, delta=0, path='best_model_molsub.pt'):
        """
        Inicjalizuje obiekt wczesnego zatrzymywania.

        Parametry:
            patience (int): Liczba epok, przez które czekamy na poprawę.
            verbose (bool): Czy wyświetlać dodatkowe informacje.
            delta (float): Minimalna różnica uznawana za poprawę.
            path (str): Ścieżka do zapisu najlepszego modelu.

        Zwraca:
            None
        """
        self.patience = patience
        self.verbose = verbose
        self.counter = 0
        self.best_score = None
        self.early_stop = False
        self.delta = delta
        self.path = path
        self.best_val_loss = np.Inf

    def __call__(self, val_loss, model):
        """
        Sprawdza, czy nastąpiła poprawa metryki. Jeżeli nie ma poprawy, zwiększa licznik.
        Po przekroczeniu 'patience' przerywa trenowanie.

        Parametry:
            val_loss (float): Strata walidacyjna.
            model (nn.Module): Trenowany model.

        Zwraca:
            None
        """
        score = -val_loss
        if self.best_score is None:
            self.best_score = score
            self.save_checkpoint(val_loss, model)
        elif score < self.best_score + self.delta:
            self.counter += 1
            if self.verbose:
                print(f'EarlyStopping counter: {self.counter} out of {self.patience}')
            if self.counter >= self.patience:
                self.early_stop = True
        else:
            self.best_score = score
            self.save_checkpoint(val_loss, model)
            self.counter = 0

    def save_checkpoint(self, val_loss, model):
        """
        Zapisuje stan modelu, jeśli osiągnięto lepszy wynik walidacyjny.

        Parametry:
            val_loss (float): Bieżąca strata walidacyjna.
            model (nn.Module): Model, który jest zapisywany.

        Zwraca:
            None
        """
        if self.verbose:
            print(f'Validation loss decreased ({self.best_val_loss:.6f} --> {val_loss:.6f}).  Saving model ...')
        torch.save(model.state_dict(), self.path)
        self.best_val_loss = val_loss


def train_network(model, optimizer, criterion, train_loader, val_loader, device, scheduler, num_epochs=120,
                  early_stopping_patience=7, unfreeze_epoch=10, output_excel_path='training_metrics.xlsx'):
    """
    Funkcja trenująca model Vision Transformer z wczesnym zatrzymywaniem i możliwością odblokowania (unfreeze)
    warstw ViT po zadanej liczbie epok.

    Parametry:
        model (nn.Module): Trenowany model.
        optimizer (torch.optim.Optimizer): Optymalizator.
        criterion (nn.Module): Funkcja straty.
        train_loader (DataLoader): Loader zbioru treningowego.
        val_loader (DataLoader): Loader zbioru walidacyjnego.
        device (torch.device): Urządzenie (CPU/GPU).
        scheduler (torch.optim.lr_scheduler): Scheduler do regulacji tempa uczenia.
        num_epochs (int): Liczba epok treningowych.
        early_stopping_patience (int): Liczba epok bez poprawy metryki, po której zatrzymujemy trening.
        unfreeze_epoch (int): Epoka, po której odblokowujemy warstwy ViT.
        output_excel_path (str): Ścieżka do pliku Excel, do którego zapisywane są metryki.

    Zwraca:
        tuple: (model, epoch_metrics)
               model (nn.Module): Najlepszy wytrenowany model (załadowany po wczesnym zatrzymaniu).
               epoch_metrics (list[dict]): Lista słowników z metrykami dla każdej epoki.
    """
    early_stopping = EarlyStopping(patience=early_stopping_patience, verbose=True, path='best_model_molsub.pt')
    epoch_metrics = []

    for epoch in range(num_epochs):
        if early_stopping.early_stop:
            print("Early stopping triggered.")
            break

        model.train()
        train_loss = 0.0
        correct = 0
        total = 0

        with tqdm(train_loader, desc=f"Epoch {epoch + 1}/{num_epochs} [Training]") as pbar:
            for images_batch, labels_batch in pbar:
                images_batch, labels_batch = images_batch.to(device), labels_batch.to(device)
                optimizer.zero_grad()
                outputs = model(images_batch)
                loss = criterion(outputs, labels_batch)
                loss.backward()
                optimizer.step()

                train_loss += loss.item() * labels_batch.size(0)
                _, predicted = outputs.max(1)
                correct += predicted.eq(labels_batch).sum().item()
                total += labels_batch.size(0)
                pbar.set_postfix(loss=f"{train_loss / total:.4f}", accuracy=f"{correct / total:.4f}")

        train_acc = correct / total
        train_loss /= total
        print(f'Train Loss: {train_loss:.3f} | Train Accuracy: {train_acc:.3f}')

        val_loss, val_acc, val_class_metrics = validate_network(model, val_loader, device, criterion)

        epoch_data = {
            'Epoch': epoch + 1,
            'Train Loss': train_loss,
            'Train Accuracy': train_acc,
            'Validation Loss': val_loss,
            'Validation Accuracy': val_acc,
            'Precision Subtype 0': val_class_metrics.get("Subtype 0", {}).get('precision', 0.0),
            'Precision Subtype 1': val_class_metrics.get("Subtype 1", {}).get('precision', 0.0),
            'Precision Subtype 2': val_class_metrics.get("Subtype 2", {}).get('precision', 0.0),
            'Precision Subtype 3': val_class_metrics.get("Subtype 3", {}).get('precision', 0.0),
            'Recall Subtype 0': val_class_metrics.get("Subtype 0", {}).get('recall', 0.0),
            'Recall Subtype 1': val_class_metrics.get("Subtype 1", {}).get('recall', 0.0),
            'Recall Subtype 2': val_class_metrics.get("Subtype 2", {}).get('recall', 0.0),
            'Recall Subtype 3': val_class_metrics.get("Subtype 3", {}).get('recall', 0.0),
        }
        epoch_metrics.append(epoch_data)

        print(f"Validation: Loss: {val_loss:.4f}, Acc: {val_acc:.4f}")
        print("Dokładność per-klasa (walidacja):")
        for cls, metrics in val_class_metrics.items():
            print(f"{cls}: Precision: {metrics['precision']:.2f}, Recall: {metrics['recall']:.2f}")

        df_epoch = pd.DataFrame([epoch_data])
        append_df_to_excel(output_excel_path, df_epoch, sheet_name='Training_Validation_Metrics')

        if isinstance(scheduler, torch.optim.lr_scheduler.ReduceLROnPlateau):
            scheduler.step(val_loss)
        else:
            scheduler.step()

        early_stopping(val_loss, model)
        if early_stopping.early_stop:
            print("Early stopping")
            break

        if epoch + 1 == unfreeze_epoch:
            print("Odblokowywanie warstw ViT do trenowania.")
            for param in model.vit.parameters():
                param.requires_grad = True
            optimizer = torch.optim.Adam(model.parameters(), lr=5e-8, weight_decay=1e-4)

    model.load_state_dict(torch.load('best_model_molsub.pt'))
    return model, epoch_metrics


def visualize_sample_images(original_image, augmented_image):
    """
    Wizualizuje oryginalny i zaugmentowany obraz obok siebie.

    Parametry:
        original_image (np.ndarray): Oryginalny obraz w formacie [W, H].
        augmented_image (np.ndarray): Zaugmentowany obraz w formacie [W, H].

    Zwraca:
        None
    """
    plt.figure(figsize=(6, 3))
    plt.subplot(1, 2, 1)
    plt.imshow(original_image, cmap='gray')
    plt.title('Oryginalny obraz')
    plt.axis('off')

    plt.subplot(1, 2, 2)
    plt.imshow(augmented_image, cmap='gray')
    plt.title('Obraz po augmentacji')
    plt.axis('off')

    plt.tight_layout()
    plt.show()


from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows


def append_df_to_excel(filename, df, sheet_name='Training_Validation_Metrics'):
    """
    Dopisuje dany DataFrame do istniejącego arkusza w pliku Excel lub tworzy nowy plik/arkusz,
    jeśli nie istnieje.

    Parametry:
        filename (str): Ścieżka do pliku Excel.
        df (pd.DataFrame): DataFrame, który chcemy dopisać.
        sheet_name (str): Nazwa arkusza w Excelu, do którego dopisujemy dane.
    """
    if not os.path.isfile(filename):
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)
    else:
        try:
            wb = load_workbook(filename)
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
            else:
                ws = wb.create_sheet(sheet_name)
            for r in dataframe_to_rows(df, index=False, header=False):
                ws.append(r)
            wb.save(filename)
        except Exception as e:
            print(f"Error while appending to Excel: {e}")


def main():
    """
    Funkcja główna skryptu.
    1. Przetwarza dane (ROI, wczytuje pliki DICOM, tworzy przetworzone obrazy).
    2. Ładuje przetworzone dane i dzieli je na zbiory: treningowy, walidacyjny i testowy.
    3. Wykonuje augmentację na zbiorze treningowym, tworząc dodatkowe przykłady.
    4. Tworzy obiekty Dataset i DataLoader dla każdego zbioru.
    5. Definiuje i trenuje model Vision Transformer z wczesnym zatrzymaniem (EarlyStopping).
    6. Waliduje i testuje model, zapisując metryki do plików Excel.

    Zwraca:
        None
    """
    processed_folder = r"C:\Users\a\Desktop\MRI\processed_images"
    train_folder = os.path.join(processed_folder, "trainMOLSUB")
    val_folder = os.path.join(processed_folder, "validationMOLSUB")
    test_folder = os.path.join(processed_folder, "testMOLSUB")
    output_excel_path = r"C:\Users\a\Desktop\MRI\Model26MOLSUB_SPRAWDZAM.xlsx"
    os.makedirs(processed_folder, exist_ok=True)

    processor = MRIDataProcessor(
        base_path=r"C:\Users\a\Desktop\MRI\manifest-1734817926430\Duke-Breast-Cancer-MRI",
        roi_excel_path=r"C:\Users\a\Desktop\MRI\Annotation_Boxes.xlsx",
        clinical_excel_path=r"C:\Users\a\Desktop\MRI\Clinical_and_Other_Features.xlsx",
        processed_folder=processed_folder,
        output_excel_path=output_excel_path,
    )

    labels_csv_path = os.path.join(processed_folder, "labels.csv")
    if not os.path.exists(labels_csv_path):
        print("Labels file not found. Starting data processing...")
        processor.process_and_save_data()
    else:
        print("Labels file found. Proceeding to next steps.")

    images, labels, class_counts = processor.load_filtered_data()
    labels = np.array(labels)
    assert labels.max() <= 3, f"Found label {labels.max()} which is out of bounds."
    assert labels.min() >= 0, f"Found label {labels.min()} which is out of bounds."

    print("Unique labels in the dataset after mapping:", np.unique(labels))

    x_train, x_val, x_test, y_train, y_val, y_test = processor.ensure_train_test_split(
        images, labels, test_size=0.1, val_size=0.1
    )

    print("\nClass distribution in Validation Set:")
    unique_val, counts_val = np.unique(y_val, return_counts=True)
    for cls, count in zip(unique_val, counts_val):
        print(f"Subtype {cls}: {count} samples")

    print("\nClass distribution in Test Set:")
    unique_test, counts_test = np.unique(y_test, return_counts=True)
    for cls, count in zip(unique_test, counts_test):
        print(f"Subtype {cls}: {count} samples")

    print("\nSaving original training data to train folder...")
    original_train_labels_csv_path = os.path.join(train_folder, "labels_train.csv")
    processor.save_split_data(x_train, y_train, train_folder, original_train_labels_csv_path)

    print("\nStarting training set augmentation to balance classes...")
    os.makedirs(train_folder, exist_ok=True)

    max_count = max([class_counts.get(cls, 0) for cls in range(4)])
    print(f"Maximum number of samples in a class: {max_count}")
    augmentation_info = {"Set": [], "Class": [], "Before_Augmentation": [], "After_Augmentation": []}

    for cls in range(4):
        current_count = class_counts.get(cls, 0)
        required_aug = max_count - current_count
        augmentation_info["Set"].append("Train")
        augmentation_info["Class"].append(f"Subtype {cls}")
        augmentation_info["Before_Augmentation"].append(current_count)

        if required_aug > 0:
            print(f"Augmenting Subtype {cls} with {required_aug} new samples.")
            class_images = [img for img, lbl in zip(x_train, y_train) if lbl == cls]
            if len(class_images) == 0:
                print(f"No images found for Subtype {cls}. Skipping augmentation.")
                augmentation_info["After_Augmentation"].append(0)
                continue

            transform = A.Compose([
                A.HorizontalFlip(p=0.5),
                A.VerticalFlip(p=0.5),
                A.RandomBrightnessContrast(p=0.3),
                A.Rotate(limit=30, p=0.3),
                A.ShiftScaleRotate(shift_limit=0.1, scale_limit=0.2, rotate_limit=15, p=0.5),
                A.GaussianBlur(blur_limit=(3, 7), p=0.3),
            ])

            for i in range(required_aug):
                img = class_images[np.random.randint(0, len(class_images))]
                augmented = transform(image=img)
                augmented_image = augmented["image"]

                filename = f"aug_subtype_{cls}_{i}.png"
                save_path = os.path.join(train_folder, filename)
                augmented_image_uint8 = (augmented_image * 255).astype(np.uint8)
                cv2.imwrite(save_path, augmented_image_uint8)

                augmented_labels = {'filename': filename, 'subtype': cls}
                augmented_labels_df = pd.DataFrame([augmented_labels])
                csv_path = os.path.join(train_folder, "labels_augmented.csv")
                if not os.path.exists(csv_path):
                    augmented_labels_df.to_csv(csv_path, index=False, header=True)
                else:
                    augmented_labels_df.to_csv(csv_path, mode='a', index=False, header=False)

            augmentation_info["After_Augmentation"].append(required_aug)
        else:
            print(f"No augmentation needed for Subtype {cls}.")
            augmentation_info["After_Augmentation"].append(0)

    if os.path.exists(os.path.join(train_folder, "labels_augmented.csv")):
        augmented_labels_df = pd.read_csv(os.path.join(train_folder, "labels_augmented.csv"))
        augmented_counts = augmented_labels_df['subtype'].value_counts().to_dict()
    else:
        augmented_counts = {}

    train_counts_after = {}
    for cls in range(4):
        original = class_counts.get(cls, 0)
        augmented = augmented_counts.get(cls, 0)
        train_counts_after[f"Subtype {cls}"] = original + augmented

    sheet1_data = {
        'Set': [],
        'Class': [],
        'Count': []
    }

    for cls, count in class_counts.items():
        sheet1_data['Set'].append('Train (Before Augmentation)')
        sheet1_data['Class'].append(f"Subtype {cls}")
        sheet1_data['Count'].append(count)

    for cls in range(4):
        count = train_counts_after.get(f"Subtype {cls}", 0)
        sheet1_data['Set'].append('Train (After Augmentation)')
        sheet1_data['Class'].append(f"Subtype {cls}")
        sheet1_data['Count'].append(count)

    for cls, count in zip(unique_val, counts_val):
        sheet1_data['Set'].append('Validation')
        sheet1_data['Class'].append(f"Subtype {cls}")
        sheet1_data['Count'].append(count)

    for cls, count in zip(unique_test, counts_test):
        sheet1_data['Set'].append('Test')
        sheet1_data['Class'].append(f"Subtype {cls}")
        sheet1_data['Count'].append(count)

    sheet1_df = pd.DataFrame(sheet1_data)

    print("\nAugmentation Info:")
    for set_type, cls, before, after in zip(augmentation_info["Set"], augmentation_info["Class"],
                                            augmentation_info["Before_Augmentation"],
                                            augmentation_info["After_Augmentation"]):
        print(f"{set_type} - {cls}: Before Augmentation: {before}, After Augmentation: {after}")

    print("\nClass distribution in Training Set before augmentation:")
    for cls in range(4):
        count = class_counts.get(cls, 0)
        print(f"Subtype {cls}: {count} samples")

    print("\nClass distribution in Training Set after augmentation:")
    for cls in range(4):
        count = train_counts_after.get(f"Subtype {cls}", 0)
        print(f"Subtype {cls}: {count} samples")

    print("\nSaving validation data to validation folder...")
    labels_val_csv_path = os.path.join(val_folder, "labels_val.csv")
    processor.save_split_data(x_val, y_val, val_folder, labels_val_csv_path)

    print("\nSaving test data to test folder...")
    labels_test_csv_path = os.path.join(test_folder, "labels_test.csv")
    processor.save_split_data(x_test, y_test, test_folder, labels_test_csv_path)

    print("\nLoading original training data...")
    original_train_dataset = BreastCancerDataset(
        images_folder=train_folder,
        labels_csv=original_train_labels_csv_path,
        augment=False,
        mean=processor.calculate_mean_std(images)[0],
        std=processor.calculate_mean_std(images)[1]
    )

    augmented_labels_path = os.path.join(train_folder, "labels_augmented.csv")
    if os.path.exists(augmented_labels_path):
        augmented_train_dataset = BreastCancerDataset(
            images_folder=train_folder,
            labels_csv=augmented_labels_path,
            augment=False,
            mean=processor.calculate_mean_std(images)[0],
            std=processor.calculate_mean_std(images)[1]
        )
    else:
        print("No augmented data found. Using original training data only.")
        augmented_train_dataset = None

    if augmented_train_dataset:
        combined_train_dataset = ConcatDataset([original_train_dataset, augmented_train_dataset])
    else:
        combined_train_dataset = original_train_dataset

    all_train_labels = []
    for _, label in combined_train_dataset:
        all_train_labels.append(label.item())
    all_train_labels = np.array(all_train_labels)
    unique_train_labels = np.unique(all_train_labels)
    print(f"\nUnikalne etykiety w zbiorze treningowym: {unique_train_labels}")
    if len(unique_train_labels) < 4:
        print("WARNING: Not all classes are present in the training set!")
    else:
        print("All classes are present in the training set.")

    class_weights = compute_class_weight("balanced", classes=np.unique(all_train_labels), y=all_train_labels)
    class_weights = torch.FloatTensor(class_weights).to(torch.device("cuda" if torch.cuda.is_available() else "cpu"))
    sample_weights = class_weights[all_train_labels]
    sampler = WeightedRandomSampler(weights=sample_weights, num_samples=len(sample_weights), replacement=True)
    criterion = FocalLoss(alpha=class_weights, gamma=3, reduction="mean")

    device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
    print(f"\nUsing device: {device}")

    num_classes = 4
    model = VisionTransformerModel(num_classes=num_classes).to(device)
    optimizer = torch.optim.Adam(model.parameters(), lr=1e-7, weight_decay=1e-4)
    scheduler = torch.optim.lr_scheduler.ReduceLROnPlateau(optimizer, mode='min', factor=0.1, patience=3, verbose=True)

    for param in model.vit.parameters():
        param.requires_grad = False

    print("\nCreating Datasets and Dataloaders...")
    train_loader = DataLoader(combined_train_dataset, batch_size=16, sampler=sampler)
    val_dataset = BreastCancerDataset(
        images_folder=val_folder,
        labels_csv=os.path.join(val_folder, "labels_val.csv"),
        augment=False,
        mean=processor.calculate_mean_std(images)[0],
        std=processor.calculate_mean_std(images)[1]
    )
    test_dataset = BreastCancerDataset(
        images_folder=test_folder,
        labels_csv=os.path.join(test_folder, "labels_test.csv"),
        augment=False,
        mean=processor.calculate_mean_std(images)[0],
        std=processor.calculate_mean_std(images)[1]
    )
    val_loader = DataLoader(val_dataset, batch_size=16, shuffle=False)
    test_loader = DataLoader(test_dataset, batch_size=16, shuffle=False)

    def check_classes_present(labels_array, split_name):
        """
        Sprawdza i wyświetla klasy obecne w danym zbiorze danych.

        Parametry:
            labels_array (np.ndarray): Tablica etykiet.
            split_name (str): Nazwa zbioru (np. 'Treningowy', 'Walidacyjny', 'Testowy').

        Zwraca:
            None
        """
        unique_labels_check = np.unique(labels_array)
        print(f"\nUnikalne etykiety w zbiorze {split_name}: {unique_labels_check}")
        if len(unique_labels_check) < 4:
            print(f"WARNING: Nie wszystkie klasy są obecne w zbiorze {split_name}!")
        else:
            print(f"All classes are present in the {split_name} set.")

    check_classes_present(y_train, "Treningowy")
    check_classes_present(y_val, "Walidacyjny")
    check_classes_present(y_test, "Testowy")

    print("\nStarting model training...")
    trained_model, epoch_metrics = train_network(
        model, optimizer, criterion, train_loader, val_loader,
        device, scheduler,
        num_epochs=120,
        early_stopping_patience=5,
        unfreeze_epoch=10,
        output_excel_path=output_excel_path
    )

    print("\nEvaluating on the test set:")
    test_loss, test_acc, test_class_metrics = validate_network(trained_model, test_loader, device, criterion)
    sheet3_data = {
        'Test Loss': [test_loss],
        'Test Accuracy': [test_acc],
        'Precision Subtype 0': [test_class_metrics.get("Subtype 0", {}).get('precision', 0.0)],
        'Precision Subtype 1': [test_class_metrics.get("Subtype 1", {}).get('precision', 0.0)],
        'Precision Subtype 2': [test_class_metrics.get("Subtype 2", {}).get('precision', 0.0)],
        'Precision Subtype 3': [test_class_metrics.get("Subtype 3", {}).get('precision', 0.0)],
        'Recall Subtype 0': [test_class_metrics.get("Subtype 0", {}).get('recall', 0.0)],
        'Recall Subtype 1': [test_class_metrics.get("Subtype 1", {}).get('recall', 0.0)],
        'Recall Subtype 2': [test_class_metrics.get("Subtype 2", {}).get('recall', 0.0)],
        'Recall Subtype 3': [test_class_metrics.get("Subtype 3", {}).get('recall', 0.0)],
    }
    sheet3_df = pd.DataFrame(sheet3_data)

    with pd.ExcelWriter(output_excel_path, engine='openpyxl') as writer:
        if os.path.exists(output_excel_path):
            book = load_workbook(output_excel_path)
            writer.book = book
            writer.sheets = {ws.title: ws for ws in book.worksheets}
        sheet1_df.to_excel(writer, sheet_name='Class_Distribution', index=False)
        sheet3_df.to_excel(writer, sheet_name='Test_Results', index=False)

    print("\nAll data has been written to the Excel file.")

    print("\nFinal Class Distributions:")
    print("Validation Set:")
    for cls, count in zip(unique_val, counts_val):
        print(f"Subtype {cls}: {count} samples")

    print("\nTest Set:")
    for cls, count in zip(unique_test, counts_test):
        print(f"Subtype {cls}: {count} samples")


if __name__ == "__main__":
    main()
