import os
import glob
from collections import defaultdict

import pandas as pd
import numpy as np
import cv2
import torch
import torch.nn as nn
from torch.utils.data import Dataset, DataLoader, WeightedRandomSampler, Subset, ConcatDataset
from transformers import ViTModel
from sklearn.model_selection import train_test_split
from sklearn.metrics import classification_report, confusion_matrix
from sklearn.utils.class_weight import compute_class_weight
from tqdm import tqdm
import albumentations as A
from albumentations.pytorch import ToTensorV2
import matplotlib.pyplot as plt
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import pydicom
import seaborn as sns
import warnings
from sklearn.exceptions import UndefinedMetricWarning

warnings.filterwarnings("ignore", category=UndefinedMetricWarning)
warnings.filterwarnings("ignore", category=UserWarning)


class MammographyDataProcessor:
    """
    Klasa służąca do przetwarzania danych mammograficznych.

    Atrybuty:
        PATHOLOGY_MAPPING (dict): Mapowanie nazw patologii na wartości liczbowe (etykiety).
        train_csv_path (str): Ścieżka do pliku CSV zawierającego dane treningowe.
        test_csv_path (str): Ścieżka do pliku CSV zawierającego dane testowe.
        dicom_base_path (str): Ścieżka bazowa do struktury folderów z plikami DICOM.
        output_train_folder (str): Ścieżka do folderu, w którym zostaną zapisane obrazy PNG (zbiór treningowy).
        output_test_folder (str): Ścieżka do folderu, w którym zostaną zapisane obrazy PNG (zbiór testowy).
    """

    PATHOLOGY_MAPPING = {
        'MALIGNANT': 0,
        'BENIGN': 1,
        'BENIGN_WITHOUT_CALLBACK': 1
    }

    def __init__(
        self,
        train_csv_path: str,
        test_csv_path: str,
        dicom_base_path: str,
        output_train_folder: str,
        output_test_folder: str
    ):
        """
        Inicjalizuje klasę MammographyDataProcessor, tworzy foldery wyjściowe, jeśli nie istnieją.

        Parametry:
            train_csv_path (str): Ścieżka do pliku CSV z danymi treningowymi.
            test_csv_path (str): Ścieżka do pliku CSV z danymi testowymi.
            dicom_base_path (str): Ścieżka bazowa do plików DICOM.
            output_train_folder (str): Ścieżka do folderu na obrazy PNG (trening).
            output_test_folder (str): Ścieżka do folderu na obrazy PNG (test).

        Zwraca:
            None
        """
        self.train_csv_path = train_csv_path
        self.test_csv_path = test_csv_path
        self.dicom_base_path = dicom_base_path
        self.output_train_folder = output_train_folder
        self.output_test_folder = output_test_folder

        os.makedirs(self.output_train_folder, exist_ok=True)
        os.makedirs(self.output_test_folder, exist_ok=True)

    def map_pathology(self, pathology: str) -> int:
        """
        Mapuje etykietę patologii (string) na wartość numeryczną (int).

        Parametry:
            pathology (str): Nazwa patologii.

        Zwraca:
            int: Wartość liczbowa etykiety (0 lub 1). Zwraca -1, jeśli etykieta nie jest rozpoznana.
        """
        return self.PATHOLOGY_MAPPING.get(pathology.upper(), -1)

    def find_first_dicom(self, patient_id: str, laterality: str, view: str, is_train: bool = True) -> str:
        """
        Znajduje plik DICOM z najmniejszą ilością czarnych pikseli dla danego pacjenta
        na podstawie nazwy folderu i struktury plików.

        Parametry:
            patient_id (str): Identyfikator pacjenta.
            laterality (str): Strona (LEFT lub RIGHT).
            view (str): Rodzaj projekcji (CC lub MLO).
            is_train (bool): Określa, czy szukamy w zbiorze treningowym (True) czy testowym (False).

        Zwraca:
            str: Ścieżka do wybranego pliku DICOM.
                 Jeśli nie odnaleziono pasującego pliku, zwraca None.
        """
        prefix = 'Mass-Training' if is_train else 'Mass-Test'
        search_pattern = f"{prefix}_{patient_id}_{laterality}_{view}_1"
        folder_path = os.path.join(self.dicom_base_path, search_pattern)
        print(f"Sprawdzam folder: {folder_path}")

        if not os.path.exists(folder_path):
            print(f"Nie znaleziono folderu DICOM dla pacjenta {patient_id}, późniejszość {laterality}, widok {view}.")
            return None

        dicom_files = glob.glob(os.path.join(folder_path, '**', '*.dcm'), recursive=True)
        print(f"Znaleziono pliki DICOM: {dicom_files}")

        if not dicom_files:
            print(f"Nie znaleziono plików DICOM w folderze {folder_path}.")
            return None

        def calculate_black_ratio(dicom_path: str) -> float:
            """
            Oblicza stosunek liczby czarnych pikseli (poniżej progu 10) do ogólnej liczby pikseli w obrazie DICOM.

            Parametry:
                dicom_path (str): Ścieżka do pliku DICOM.

            Zwraca:
                float: Wartość w przedziale [0,1], oznaczająca udział czarnych pikseli w obrazie.
            """
            try:
                dicom = pydicom.dcmread(dicom_path)
                image = dicom.pixel_array
                normalized_image = cv2.normalize(image, None, 0, 255, cv2.NORM_MINMAX).astype(np.uint8)
                black_pixels = np.sum(normalized_image < 10)
                total_pixels = normalized_image.size
                return black_pixels / total_pixels
            except Exception as e:
                print(f"Nie udało się przetworzyć pliku {dicom_path}: {e}")
                return 1

        black_ratios = {dicom_path: calculate_black_ratio(dicom_path) for dicom_path in dicom_files}
        best_dicom = min(black_ratios, key=black_ratios.get)
        print(
            f"Wybrano plik DICOM z najmniejszą ilością czerni: {best_dicom} "
            f"(czarność: {black_ratios[best_dicom]:.2%})"
        )
        return best_dicom

    def convert_dicom_to_png(self, dicom_path: str) -> np.ndarray:
        """
        Konwertuje plik DICOM do obrazu PNG, stosując normalizację i CLAHE.

        Parametry:
            dicom_path (str): Ścieżka do pliku DICOM.

        Zwraca:
            np.ndarray: Obraz w formacie RGB. Zwraca None w przypadku błędu.
        """
        try:
            dicom = pydicom.dcmread(dicom_path)
            image = dicom.pixel_array
            image = cv2.normalize(image, None, 0, 255, cv2.NORM_MINMAX).astype('uint8')
            clahe = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(8, 8))
            image = clahe.apply(image)
            image_rgb = cv2.cvtColor(image, cv2.COLOR_GRAY2RGB)
            return image_rgb
        except Exception as e:
            print(f"Error processing DICOM file {dicom_path}: {e}")
            return None

    def process_csv(self, csv_path: str, output_folder: str, label_csv_path: str, is_train: bool = True) -> None:
        """
        Przetwarza plik CSV (treningowy lub testowy), konwertuje pasujące pliki DICOM do formatu PNG
        i zapisuje etykiety w pliku CSV.

        Parametry:
            csv_path (str): Ścieżka do pliku CSV z danymi (treningowym lub testowym).
            output_folder (str): Folder wyjściowy na obrazy PNG.
            label_csv_path (str): Ścieżka do wyjściowego pliku CSV z etykietami.
            is_train (bool): Czy zbiór jest treningowy (True) czy testowy (False).

        Zwraca:
            None
        """
        df = pd.read_csv(csv_path)
        print(f"Przetwarzanie pliku: {csv_path}")
        print("Dostępne kolumny:", df.columns.tolist())

        pathology_col = 'pathology'
        patient_id_col = 'patient_id'
        laterality_col = 'left or right breast'
        view_col = 'image view'
        required_columns = [pathology_col, patient_id_col, laterality_col, view_col]
        missing_columns = [col for col in required_columns if col not in df.columns]

        if missing_columns:
            raise ValueError(f"Brakujące kolumny w pliku CSV {csv_path}: {missing_columns}")

        labels = []
        processed_count = 0

        for idx, row in tqdm(df.iterrows(), total=df.shape[0], desc=f"Przetwarzanie {os.path.basename(csv_path)}"):
            pathology = row[pathology_col]
            mapped_label = self.map_pathology(pathology)
            if mapped_label == -1:
                print(f"Nieznana etykieta patologii '{pathology}' dla indeksu {idx}. Pomijanie...")
                continue

            patient_id = row[patient_id_col]
            laterality = row[laterality_col].strip().upper()
            view = row[view_col].strip().upper()
            png_filename = f"{patient_id}_{laterality}_{view}_{idx}.png"
            png_path = os.path.join(output_folder, png_filename)

            if os.path.exists(png_path):
                print(f"Plik {png_filename} już istnieje. Pomijanie przetwarzania...")
                labels.append({'filename': png_filename, 'pathology': mapped_label})
                processed_count += 1
                continue

            dicom_path = self.find_first_dicom(patient_id, laterality, view, is_train)
            if not dicom_path:
                print(
                    f"Nie znaleziono pliku DICOM dla pacjenta {patient_id}, "
                    f"późniejszość {laterality}, widok {view}."
                )
                continue

            image = self.convert_dicom_to_png(dicom_path)
            if image is None:
                print(f"Konwersja DICOM do PNG nie powiodła się dla pliku {dicom_path}.")
                continue

            cv2.imwrite(png_path, image)
            labels.append({'filename': png_filename, 'pathology': mapped_label})
            processed_count += 1

        labels_df = pd.DataFrame(labels)
        labels_df.to_csv(label_csv_path, index=False)
        print(f"Przetworzono {processed_count} obrazów z pliku {os.path.basename(csv_path)}.")
        print(f"Etykiety zapisano do {label_csv_path}.\n")

    def run(self) -> None:
        """
        Uruchamia proces przetwarzania zbiorów treningowego i testowego.
        Jeśli już istnieją odpowiednie pliki CSV z etykietami, proces zostanie pominięty.

        Zwraca:
            None
        """
        train_label_csv = os.path.join(self.output_train_folder, "labels.csv")
        test_label_csv = os.path.join(self.output_test_folder, "labels.csv")

        if not os.path.exists(train_label_csv):
            print("Przetwarzanie zbioru treningowego...")
            self.process_csv(
                csv_path=self.train_csv_path,
                output_folder=self.output_train_folder,
                label_csv_path=train_label_csv,
                is_train=True
            )
        else:
            print(
                f"Plik {train_label_csv} już istnieje. "
                f"Pomijanie przetwarzania zbioru treningowego."
            )

        if not os.path.exists(test_label_csv):
            print("Przetwarzanie zbioru testowego...")
            self.process_csv(
                csv_path=self.test_csv_path,
                output_folder=self.output_test_folder,
                label_csv_path=test_label_csv,
                is_train=False
            )
        else:
            print(
                f"Plik {test_label_csv} już istnieje. "
                f"Pomijanie przetwarzania zbioru testowego."
            )
        print("Przetwarzanie zakończone.")


class MammographyDataset(Dataset):
    """
    Klasa Dataset do wczytywania obrazów mammograficznych i etykiet z pliku CSV,
    opcjonalnie z augmentacją Albumentations.
    """

    def __init__(
        self,
        images_folder: str,
        labels_csv: str,
        augment: bool = False,
        mean: np.ndarray = None,
        std: np.ndarray = None
    ):
        """
        Inicjalizuje dataset, wczytując ścieżki do obrazów (PNG) i odpowiadające etykiety.

        Parametry:
            images_folder (str): Ścieżka do folderu z obrazami PNG.
            labels_csv (str): Ścieżka do pliku CSV zawierającego nazwy plików i etykiety.
            augment (bool): Czy zastosować augmentację.
            mean (np.ndarray): Średnia dla normalizacji (3 wartości, dla każdego kanału).
            std (np.ndarray): Odchylenie standardowe dla normalizacji (3 wartości).

        Zwraca:
            None
        """
        self.images_folder = images_folder
        self.labels_df = pd.read_csv(labels_csv)
        self.labels = self.labels_df['pathology'].astype(int).values
        self.filenames = self.labels_df['filename'].values
        self.augment = augment
        self.mean = mean
        self.std = std
        self.augment_counters = defaultdict(int)

        if mean is None or std is None:
            raise ValueError("Musisz podać mean i std do normalizacji obrazów.")

        if augment:
            self.transforms = A.Compose([
                A.Resize(224, 224),
                A.HorizontalFlip(p=0.3),
                A.GaussianBlur(blur_limit=(3, 7), p=0.4),
                A.MotionBlur(blur_limit=5, p=0.4),
                A.VerticalFlip(p=0.3),
                A.RandomBrightnessContrast(p=0.4),
                A.Rotate(limit=10, p=0.1),
                A.Affine(
                    scale=(0.8, 1.2),
                    translate_percent=(0.0, 0.0),
                    rotate=(-45, 45),
                    shear=(-20, 20),
                    p=0.5
                ),
                A.Normalize(mean=mean.tolist(), std=std.tolist()),
                ToTensorV2()
            ])
        else:
            self.transforms = A.Compose([
                A.Resize(224, 224),
                A.Normalize(mean=mean.tolist(), std=std.tolist()),
                ToTensorV2()
            ])

    def __len__(self) -> int:
        """
        Zwraca liczbę próbek w zbiorze.

        Zwraca:
            int: Liczba dostępnych obrazów.
        """
        return len(self.labels)

    def __getitem__(self, idx: int):
        """
        Zwraca obraz i jego etykietę na podstawie indeksu.

        Parametry:
            idx (int): Indeks próbki.

        Zwraca:
            tuple: (torch.Tensor, torch.Tensor) — obraz i etykieta.
        """
        filename = self.filenames[idx]
        image_path = os.path.join(self.images_folder, filename)
        image = cv2.imread(image_path, cv2.IMREAD_GRAYSCALE)
        if image is None:
            raise FileNotFoundError(f"Unable to read image: {image_path}")

        image = np.dstack([image] * 3).astype(np.float32) / 255.0

        transformed = self.transforms(image=image)
        image = transformed["image"]
        label = torch.tensor(self.labels[idx], dtype=torch.long)
        return image, label


class VisionTransformerModel(nn.Module):
    """
    Klasa modelu Vision Transformer (ViT) z dodatkową warstwą klasyfikacji.
    """

    def __init__(self, num_classes: int):
        """
        Inicjalizuje wstępnie wytrenowany model ViT oraz dodaje sekwencję warstw klasyfikujących.

        Parametry:
            num_classes (int): Liczba klas do przewidywania.

        Zwraca:
            None
        """
        super(VisionTransformerModel, self).__init__()
        self.vit = ViTModel.from_pretrained("google/vit-base-patch16-224")
        vit_hidden_size = self.vit.config.hidden_size
        self.classifier = nn.Sequential(
            nn.Linear(vit_hidden_size, 256),
            nn.LeakyReLU(0.1),
            nn.Dropout(0.4),
            nn.Linear(256, 128),
            nn.LeakyReLU(0.1),
            nn.Dropout(0.3),
            nn.Linear(128, num_classes)
        )

    def forward(self, images: torch.Tensor) -> torch.Tensor:
        """
        Definiuje przepływ danych przez model (forward pass).

        Parametry:
            images (torch.Tensor): Tensor obrazów o kształcie (batch_size, 3, 224, 224).

        Zwraca:
            torch.Tensor: Logity (wyniki przed softmaxem) dla każdej klasy w zbiorze.
        """
        outputs = self.vit(pixel_values=images)
        features = outputs.last_hidden_state[:, 0, :]
        logits = self.classifier(features)
        return logits


class ExcelLogger:
    """
    Klasa do zapisywania i aktualizowania wyników (metryk treningu, testu itd.) w pliku Excel.
    """

    def __init__(self, output_path: str):
        """
        Inicjalizuje obiekt loggera i tworzy arkusze w pliku Excel, jeśli nie istnieją.

        Parametry:
            output_path (str): Ścieżka do pliku Excel, w którym będą zapisywane wyniki.

        Zwraca:
            None
        """
        self.output_path = output_path
        self.sheet_names = [
            "Class Distribution",
            "Augmentation Info",
            "Training and Validation",
            "Test Results",
            "Per-Class Accuracy",
            "Processed_Data"
        ]
        os.makedirs(os.path.dirname(self.output_path), exist_ok=True)

        if os.path.exists(output_path):
            self.workbook = load_workbook(output_path)
            for sheet in self.sheet_names:
                if sheet not in self.workbook.sheetnames:
                    print(f"Tworzenie brakującego arkusza: {sheet}")
                    self.workbook.create_sheet(sheet)
        else:
            self.workbook = Workbook()
            default_sheet = self.workbook.active
            self.workbook.remove(default_sheet)
            for sheet in self.sheet_names:
                self.workbook.create_sheet(sheet)
            self.save()

    def save_class_distribution(
        self,
        train_counts: dict,
        val_counts: dict,
        test_counts: dict
    ) -> None:
        """
        Zapisuje informacje o rozkładzie klas w zbiorach treningowym, walidacyjnym i testowym.

        Parametry:
            train_counts (dict): Rozkład klas w zbiorze treningowym (słownik klasa->liczność).
            val_counts (dict): Rozkład klas w zbiorze walidacyjnym.
            test_counts (dict): Rozkład klas w zbiorze testowym.

        Zwraca:
            None
        """
        df = pd.DataFrame({
            "Class": ["Grade 0", "Grade 1"],
            "Train Count": [
                train_counts.get(0, 0),
                train_counts.get(1, 0)
            ],
            "Validation Count": [
                val_counts.get(0, 0),
                val_counts.get(1, 0)
            ],
            "Test Count": [
                test_counts.get(0, 0),
                test_counts.get(1, 0)
            ],
        })
        self._write_to_sheet("Class Distribution", df)

    def save_augmentation_info(self, augmentation_info: dict) -> None:
        """
        Zapisuje informacje o augmentacjach (ile razy każda transformacja została zastosowana).

        Parametry:
            augmentation_info (dict): Słownik klucz->liczba użyć augmentacji.

        Zwraca:
            None
        """
        df = pd.DataFrame(list(augmentation_info.items()), columns=["Augmentation", "Count"])
        self._write_to_sheet("Augmentation Info", df)

    def save_epoch_metrics(
        self,
        epoch: int,
        train_loss: float,
        train_acc: float,
        val_loss: float,
        val_acc: float
    ) -> None:
        """
        Zapisuje metryki treningu i walidacji dla danej epoki.

        Parametry:
            epoch (int): Numer epoki.
            train_loss (float): Strata w zbiorze treningowym.
            train_acc (float): Dokładność w zbiorze treningowym.
            val_loss (float): Strata w zbiorze walidacyjnym.
            val_acc (float): Dokładność w zbiorze walidacyjnym.

        Zwraca:
            None
        """
        sheet_name = "Training and Validation"
        sheet = self.workbook[sheet_name]
        if sheet.max_row == 1 and sheet.max_column == 1 and sheet.cell(row=1, column=1).value is None:
            headers = ["Epoch", "Train Loss", "Train Accuracy", "Validation Loss", "Validation Accuracy"]
            sheet.append(headers)
        sheet.append([epoch, train_loss, train_acc, val_loss, val_acc])

    def save_test_results(self, test_loss: float, test_acc: float) -> None:
        """
        Zapisuje metryki osiągnięte na zbiorze testowym.

        Parametry:
            test_loss (float): Strata w zbiorze testowym.
            test_acc (float): Dokładność w zbiorze testowym.

        Zwraca:
            None
        """
        sheet_name = "Test Results"
        sheet = self.workbook[sheet_name]
        if sheet.max_row == 1 and sheet.max_column == 1 and sheet.cell(row=1, column=1).value is None:
            headers = ["Test Loss", "Test Accuracy"]
            sheet.append(headers)
        sheet.append([test_loss, test_acc])

    def save_class_accuracies(
        self,
        epoch: int,
        class_accuracies: dict,
        dataset_type: str = "Validation"
    ) -> None:
        """
        Zapisuje dokładność dla każdej klasy osobno.

        Parametry:
            epoch (int): Numer epoki.
            class_accuracies (dict): Słownik klasa->dokładność (w procentach).
            dataset_type (str): Nazwa zbioru, np. "Validation" czy "Test".

        Zwraca:
            None
        """
        sheet_name = "Per-Class Accuracy"
        sheet = self.workbook[sheet_name]
        if sheet.max_row == 1 and sheet.max_column == 1 and sheet.cell(row=1, column=1).value is None:
            headers = ["Epoch", "Dataset_Type", "Class", "Accuracy (%)"]
            sheet.append(headers)

        for cls, acc in class_accuracies.items():
            sheet.append([epoch, dataset_type, f"Grade {cls}", acc])

    def save(self) -> None:
        """
        Zapisuje zmiany w pliku Excel.

        Zwraca:
            None
        """
        self.workbook.save(self.output_path)

    def _write_to_sheet(self, sheet_name: str, dataframe: pd.DataFrame, append: bool = False) -> None:
        """
        Prywatna metoda do zapisywania DataFrame w danym arkuszu pliku Excel.

        Parametry:
            sheet_name (str): Nazwa arkusza.
            dataframe (pd.DataFrame): Ramka danych do zapisania.
            append (bool): Czy dopisywać do istniejących danych (True), czy nadpisać (False).

        Zwraca:
            None
        """
        sheet = self.workbook[sheet_name]
        if not append:
            if sheet.max_row == 1 and sheet.max_column == 1 and sheet.cell(row=1, column=1).value is None:
                for row in dataframe_to_rows(dataframe, index=False, header=True):
                    sheet.append(row)
            else:
                for row in dataframe_to_rows(dataframe, index=False, header=False):
                    sheet.append(row)
        else:
            for row in dataframe_to_rows(dataframe, index=False, header=False):
                sheet.append(row)


class FocalLoss(nn.Module):
    """
    Implementacja funkcji straty Focal Loss, przydatnej przy niezrównoważonych zbiorach.
    """

    def __init__(
        self,
        alpha: torch.Tensor = None,
        gamma: float = 3,
        reduction: str = 'mean'
    ):
        """
        Inicjalizuje obiekt FocalLoss.

        Parametry:
            alpha (torch.Tensor lub None): Współczynnik ważenia klas.
            gamma (float): Parametr ogniskujący (focusing parameter).
            reduction (str): Rodzaj redukcji ('mean', 'sum' lub 'none').

        Zwraca:
            None
        """
        super(FocalLoss, self).__init__()
        self.alpha = alpha if alpha is not None else torch.tensor(1.0)
        self.gamma = gamma
        self.reduction = reduction

    def forward(self, inputs: torch.Tensor, targets: torch.Tensor) -> torch.Tensor:
        """
        Oblicza wartość straty Focal Loss.

        Parametry:
            inputs (torch.Tensor): Logity (wyjście modelu).
            targets (torch.Tensor): Poprawne etykiety.

        Zwraca:
            torch.Tensor: Wartość straty Focal Loss.
        """
        BCE_loss = nn.CrossEntropyLoss(reduction='none')(inputs, targets)
        pt = torch.exp(-BCE_loss)
        if isinstance(self.alpha, torch.Tensor):
            alpha_t = self.alpha[targets]
        else:
            alpha_t = self.alpha
        F_loss = alpha_t * (1 - pt) ** self.gamma * BCE_loss
        if self.reduction == 'mean':
            return F_loss.mean()
        elif self.reduction == 'sum':
            return F_loss.sum()
        else:
            return F_loss


def plot_confusion_matrix(all_labels: np.ndarray, all_preds: np.ndarray, classes: list) -> None:
    """
    Rysuje macierz konfuzji dla zadanych etykiet i przewidywań.

    Parametry:
        all_labels (np.ndarray): Tablica z prawdziwymi etykietami.
        all_preds (np.ndarray): Tablica z przewidzianymi etykietami.
        classes (list): Lista nazw klas do opisania osi.

    Zwraca:
        None
    """
    cm = confusion_matrix(all_labels, all_preds)
    plt.figure(figsize=(8, 6))
    sns.heatmap(cm, annot=True, fmt='d', cmap='Blues',
                xticklabels=classes, yticklabels=classes)
    plt.xlabel('Predicted')
    plt.ylabel('True')
    plt.title('Confusion Matrix')
    plt.show()


def validate_network(
    model: nn.Module,
    val_loader: DataLoader,
    device: torch.device,
    criterion: nn.Module
) -> tuple:
    """
    Waliduje model na podanym zbiorze danych, obliczając stratę, dokładność i dokładność dla każdej klasy.

    Parametry:
        model (nn.Module): Model do walidacji.
        val_loader (DataLoader): Loader zbioru walidacyjnego lub testowego.
        device (torch.device): Urządzenie (CPU/GPU) do obliczeń.
        criterion (nn.Module): Funkcja straty.

    Zwraca:
        tuple: (val_loss, val_acc, class_accuracies)
    """
    model.eval()
    val_loss = 0.0
    correct = 0
    total = 0
    all_preds = []
    all_labels = []

    with torch.no_grad():
        for (x_val, y_val) in val_loader:
            x_val, y_val = x_val.to(device), y_val.to(device)
            outputs = model(x_val)
            loss = criterion(outputs, y_val)
            val_loss += loss.item() * y_val.size(0)
            _, predicted = outputs.max(1)
            correct += predicted.eq(y_val).sum().item()
            total += y_val.size(0)
            all_preds.append(predicted.cpu().numpy())
            all_labels.append(y_val.cpu().numpy())

    val_loss /= total
    val_acc = correct / total
    all_preds = np.concatenate(all_preds)
    all_labels = np.concatenate(all_labels)
    report = classification_report(all_labels, all_preds, target_names=["Grade 0", "Grade 1"], zero_division=0)
    print(report)
    plot_confusion_matrix(all_labels, all_preds, classes=["Grade 0", "Grade 1"])
    classes_unique = np.unique(all_labels)
    class_accuracies = {}
    for cls in classes_unique:
        cls_mask = (all_labels == cls)
        cls_acc = (all_preds[cls_mask] == cls).mean() if np.any(cls_mask) else 0.0
        class_accuracies[cls] = cls_acc * 100.0
    return val_loss, val_acc, class_accuracies


class EarlyStopping:
    """
    Klasa implementująca wczesne zatrzymanie (Early Stopping) trenowania,
    jeśli metryka walidacyjna nie poprawia się przez określoną liczbę epok.
    """

    def __init__(
        self,
        patience: int = 5,
        verbose: bool = False,
        delta: float = 0,
        path: str = 'best_model_mam5.pt'
    ):
        """
        Inicjalizuje EarlyStopping.

        Parametry:
            patience (int): Liczba epok braku poprawy, po której kończymy trenowanie.
            verbose (bool): Czy wypisywać informacje o zapisywaniu modelu.
            delta (float): Minimalna zmiana metryki uznawana za poprawę.
            path (str): Ścieżka do zapisu najlepszego modelu.

        Zwraca:
            None
        """
        self.patience = patience
        self.verbose = verbose
        self.counter = 0
        self.best_score = None
        self.early_stop = False
        self.delta = delta
        self.path = path
        self.best_val_loss = np.Inf

    def __call__(self, val_loss: float, model: nn.Module) -> None:
        """
        Główny punkt wejścia EarlyStopping, wywoływany po każdej epoce.

        Parametry:
            val_loss (float): Bieżąca wartość straty walidacyjnej.
            model (nn.Module): Aktualnie trenowany model.

        Zwraca:
            None
        """
        score = -val_loss
        if self.best_score is None:
            self.best_score = score
            self.save_checkpoint(val_loss, model)
        elif score < self.best_score + self.delta:
            self.counter += 1
            if self.verbose:
                print(f'EarlyStopping counter: {self.counter} out of {self.patience}')
            if self.counter >= self.patience:
                self.early_stop = True
        else:
            self.best_score = score
            self.save_checkpoint(val_loss, model)
            self.counter = 0

    def save_checkpoint(self, val_loss: float, model: nn.Module) -> None:
        """
        Zapisuje stan modelu, gdy osiągnięto najlepszy wynik (najniższa strata walidacyjna).

        Parametry:
            val_loss (float): Bieżąca strata walidacyjna.
            model (nn.Module): Model, który jest zapisywany.

        Zwraca:
            None
        """
        if self.verbose:
            print(f'Validation loss decreased ({self.best_val_loss:.6f} --> {val_loss:.6f}).  Saving model ...')
        torch.save(model.state_dict(), self.path)
        self.best_val_loss = val_loss


def train_network(
    model: nn.Module,
    optimizer: torch.optim.Optimizer,
    criterion: nn.Module,
    train_loader: DataLoader,
    val_loader: DataLoader,
    device: torch.device,
    scheduler,
    logger: ExcelLogger,
    num_epochs: int = 21,
    early_stopping_patience: int = 5,
    unfreeze_epoch: int = 10
) -> nn.Module:
    """
    Trenuje model VisionTransformerModel z użyciem wczesnego zatrzymania (EarlyStopping) oraz umożliwia
    odblokowanie (unfreeze) warstw ViT w zadanej epoce.

    Parametry:
        model (nn.Module): Model do trenowania.
        optimizer (torch.optim.Optimizer): Optymalizator (np. Adam).
        criterion (nn.Module): Funkcja straty (np. FocalLoss).
        train_loader (DataLoader): Loader zbioru treningowego.
        val_loader (DataLoader): Loader zbioru walidacyjnego.
        device (torch.device): Urządzenie obliczeniowe (CPU/GPU).
        scheduler: Scheduler do regulacji learning rate (np. ReduceLROnPlateau).
        logger (ExcelLogger): Obiekt do logowania wyników do pliku Excel.
        num_epochs (int): Liczba epok trenowania.
        early_stopping_patience (int): Liczba epok bez poprawy, po której przerywamy trening.
        unfreeze_epoch (int): Epoka, w której odblokowujemy warstwy ViT.

    Zwraca:
        nn.Module: Najlepszy wytrenowany model (załadowany po wystąpieniu EarlyStopping).
    """
    early_stopping = EarlyStopping(patience=early_stopping_patience, verbose=True, path='best_model_mam5.pt')

    for epoch in range(num_epochs):
        if early_stopping.early_stop:
            print("Early stopping triggered.")
            break

        model.train()
        train_loss = 0.0
        correct = 0
        total = 0

        with tqdm(train_loader, desc=f"Epoch {epoch + 1}/{num_epochs} [Training]") as pbar:
            for images_batch, labels_batch in pbar:
                images_batch, labels_batch = images_batch.to(device), labels_batch.to(device)

                optimizer.zero_grad()
                outputs = model(images_batch)
                loss = criterion(outputs, labels_batch)
                loss.backward()
                optimizer.step()

                train_loss += loss.item() * labels_batch.size(0)
                _, predicted = outputs.max(1)
                correct += predicted.eq(labels_batch).sum().item()
                total += labels_batch.size(0)

                pbar.set_postfix(loss=f"{train_loss / total:.4f}", accuracy=f"{correct / total:.4f}")

        train_acc = correct / total
        train_loss /= total
        print(f'Train Loss: {train_loss:.3f} | Train Accuracy: {train_acc:.3f}')

        val_loss, val_acc, val_class_accuracies = validate_network(model, val_loader, device, criterion)
        logger.save_epoch_metrics(epoch + 1, train_loss, train_acc, val_loss, val_acc)
        logger.save_class_accuracies(epoch + 1, val_class_accuracies, dataset_type="Validation")

        print(f"Validation: Loss: {val_loss:.4f}, Acc: {val_acc:.4f}")
        print("Dokładność per-klasa (walidacja):")
        for cls_id, acc_cls in val_class_accuracies.items():
            print(f"Grade {cls_id}: {acc_cls:.2f}%")

        if isinstance(scheduler, torch.optim.lr_scheduler.ReduceLROnPlateau):
            scheduler.step(val_loss)
        else:
            scheduler.step()

        early_stopping(val_loss, model)
        if early_stopping.early_stop:
            print("Early stopping")
            break

        if epoch + 1 == unfreeze_epoch:
            print("Odblokowywanie warstw ViT do trenowania.")
            for param in model.vit.parameters():
                param.requires_grad = True
            optimizer = torch.optim.Adam(model.parameters(), lr=1e-3, weight_decay=1e-4)

    model.load_state_dict(torch.load('best_model_mam5.pt'))
    return model


def split_test_set(
    test_folder: str,
    test_labels_csv: str,
    validation_folder: str,
    validation_csv_path: str,
    test_new_folder: str,
    test_new_csv_path: str
) -> None:
    """
    Dzieli pierwotny zbiór testowy na dwa podzbiory: walidacyjny i nowy testowy,
    zachowując równomierny rozkład klas.

    Parametry:
        test_folder (str): Ścieżka do folderu z oryginalnym zbiorem testowym.
        test_labels_csv (str): Ścieżka do pliku CSV z etykietami (oryginalny zbiór testowy).
        validation_folder (str): Ścieżka do nowego folderu walidacyjnego.
        validation_csv_path (str): Ścieżka do pliku CSV walidacyjnego.
        test_new_folder (str): Ścieżka do nowego folderu testowego.
        test_new_csv_path (str): Ścieżka do pliku CSV nowego zbioru testowego.

    Zwraca:
        None
    """
    df = pd.read_csv(test_labels_csv)
    labels = df['pathology'].values
    filenames = df['filename'].values
    os.makedirs(validation_folder, exist_ok=True)
    os.makedirs(test_new_folder, exist_ok=True)

    val_filenames = []
    val_labels = []
    test_new_filenames = []
    test_new_labels = []
    unique_labels = np.unique(labels)

    for label in unique_labels:
        label_indices = np.where(labels == label)[0]
        np.random.shuffle(label_indices)
        split_idx = len(label_indices) // 2
        val_indices = label_indices[:split_idx]
        test_new_indices = label_indices[split_idx:]
        val_filenames.extend(filenames[val_indices])
        val_labels.extend(labels[val_indices])
        test_new_filenames.extend(filenames[test_new_indices])
        test_new_labels.extend(labels[test_new_indices])

    def copy_files(filenames_list, labels_list, destination_folder, csv_path):
        """
        Kopiuje (a właściwie zapisuje ponownie) obrazy PNG do wskazanego folderu
        i tworzy plik CSV z informacjami o plikach i etykietach.
        """
        labels_data = []
        for filename, lbl in zip(filenames_list, labels_list):
            src_path = os.path.join(test_folder, filename)
            dest_path = os.path.join(destination_folder, filename)
            if not os.path.exists(dest_path):
                cv2.imwrite(dest_path, cv2.imread(src_path))
            labels_data.append({'filename': filename, 'pathology': lbl})
        labels_df = pd.DataFrame(labels_data)
        labels_df.to_csv(csv_path, index=False)

    copy_files(val_filenames, val_labels, validation_folder, validation_csv_path)
    copy_files(test_new_filenames, test_new_labels, test_new_folder, test_new_csv_path)
    print(
        f"Podział zbioru testowego zakończony. "
        f"Walidacja: {len(val_filenames)} obrazów, Test nowy: {len(test_new_filenames)} obrazów."
    )


def main() -> None:
    """
    Główna funkcja skryptu. Odpowiada za:
    1. Konwersję plików DICOM na obrazy PNG dla zbiorów treningowego i testowego.
    2. Dzielenie pierwotnego zbioru testowego na mniejszy walidacyjny i nowy testowy.
    3. Obliczenie mean i std ze zbioru treningowego.
    4. Wczytanie danych do obiektów Dataset (z augmentacją w zbiorze treningowym).
    5. Trenowanie modelu (VisionTransformerModel) z early stopping i (opcjonalnym) odblokowaniem warstw.
    6. Zapis wyników do pliku Excel i ewaluacja na nowym zbiorze testowym.

    Zwraca:
        None
    """
    train_csv = r"C:\Users\a\Desktop\mammografia\manifest-1734734641099\mass_case_description_train_set.csv"
    test_csv = r"C:\Users\a\Desktop\mammografia\manifest-1734734641099\mass_case_description_test_set.csv"
    dicom_base = r"C:\Users\a\Desktop\mammografia\manifest-1734734641099\CBIS-DDSM"
    output_train = r"C:\Users\a\Desktop\mammografia\manifest-1734734641099\MAMMOGRAFIAPNG\TRAIN"
    output_test = r"C:\Users\a\Desktop\mammografia\manifest-1734734641099\MAMMOGRAFIAPNG\TEST"

    processor = MammographyDataProcessor(
        train_csv_path=train_csv,
        test_csv_path=test_csv,
        dicom_base_path=dicom_base,
        output_train_folder=output_train,
        output_test_folder=output_test
    )
    processor.run()

    labels_train_path = os.path.join(output_train, "labels.csv")
    labels_test_path = os.path.join(output_test, "labels.csv")

    validation_folder = r"C:\Users\a\Desktop\mammografia\manifest-1734734641099\MAMMOGRAFIAPNG\VALIDATION"
    validation_csv_path = os.path.join(validation_folder, "labels_validation.csv")
    test_new_folder = r"C:\Users\a\Desktop\mammografia\manifest-1734734641099\MAMMOGRAFIAPNG\TEST_NEW"
    test_new_csv_path = os.path.join(test_new_folder, "labels_test_new.csv")

    if not (os.path.exists(validation_csv_path) and os.path.exists(test_new_csv_path)):
        print("Dzielenie zbioru testowego na walidacyjny i testowy...")
        split_test_set(
            output_test,
            labels_test_path,
            validation_folder,
            validation_csv_path,
            test_new_folder,
            test_new_csv_path
        )
    else:
        print("Podział zbioru testowego już istnieje. Pomijanie.")

    output_excel_path = r"C:\Users\a\Desktop\mammografia\Model_POPRAWKA_Mammografia_5.xlsx"
    logger = ExcelLogger(output_excel_path)

    train_dataset_for_stats = MammographyDataset(
        images_folder=output_train,
        labels_csv=labels_train_path,
        augment=False,
        mean=np.zeros(3),
        std=np.ones(3)
    )
    loader_for_stats = DataLoader(train_dataset_for_stats, batch_size=16, shuffle=False)

    def calculate_mean_std(loader: DataLoader) -> tuple:
        """
        Oblicza średnią i odchylenie standardowe dla całego zbioru,
        iterując po mini-batchach dostarczanych przez loader.

        Parametry:
            loader (DataLoader): Zbiór danych (bez augmentacji) do obliczenia statystyk.

        Zwraca:
            tuple: (mean, std) dla kanałów (R, G, B).
        """
        mean_val = 0.
        std_val = 0.
        total_samples = 0
        for images, _ in loader:
            batch_samples = images.size(0)
            images = images.view(batch_samples, images.size(1), -1)
            mean_val += images.mean(2).sum(0)
            std_val += images.std(2).sum(0)
            total_samples += batch_samples
        mean_val /= total_samples
        std_val /= total_samples
        return mean_val, std_val

    mean, std = calculate_mean_std(loader_for_stats)
    print(f"Mean: {mean}, Std: {std}")

    train_counts = defaultdict(int)
    validation_counts = defaultdict(int)
    test_new_counts = defaultdict(int)

    train_labels_df = pd.read_csv(labels_train_path)
    for label in train_labels_df['pathology']:
        train_counts[label] += 1

    validation_labels_df = pd.read_csv(validation_csv_path)
    for label in validation_labels_df['pathology']:
        validation_counts[label] += 1

    test_new_labels_df = pd.read_csv(test_new_csv_path)
    for label in test_new_labels_df['pathology']:
        test_new_counts[label] += 1

    logger.save_class_distribution(train_counts, validation_counts, test_new_counts)

    print("\nLiczba przypadków danej patologii w zbiorze treningowym:")
    for cls, count in train_counts.items():
        print(f"Grade {cls}: {count} przypadków")

    print("\nLiczba przypadków danej patologii w zbiorze walidacyjnym:")
    for cls, count in validation_counts.items():
        print(f"Grade {cls}: {count} przypadków")

    print("\nLiczba przypadków danej patologii w zbiorze testowym:")
    for cls, count in test_new_counts.items():
        print(f"Grade {cls}: {count} przypadków")

    train_dataset = MammographyDataset(
        images_folder=output_train,
        labels_csv=labels_train_path,
        augment=False,
        mean=mean,
        std=std
    )
    validation_dataset = MammographyDataset(
        images_folder=validation_folder,
        labels_csv=validation_csv_path,
        augment=False,
        mean=mean,
        std=std
    )
    test_dataset = MammographyDataset(
        images_folder=test_new_folder,
        labels_csv=test_new_csv_path,
        augment=False,
        mean=mean,
        std=std
    )

    class_counts = train_counts
    max_count = max(class_counts.values())
    class_indices = defaultdict(list)
    for idx, label in enumerate(train_dataset.labels):
        class_indices[label].append(idx)

    augmented_indices = []
    for cls, indices in class_indices.items():
        if len(indices) < max_count:
            missing = max_count - len(indices)
            additional_indices = np.random.choice(indices, size=missing, replace=True)
            augmented_indices.extend(additional_indices)
        else:
            augmented_indices.extend(indices)

    balanced_dataset = Subset(train_dataset, augmented_indices)
    combined_train_dataset = ConcatDataset([train_dataset, balanced_dataset])
    combined_labels = [train_dataset.labels[idx] for idx in augmented_indices]
    for label in combined_labels:
        train_counts[label] += 1

    logger.save_augmentation_info(train_dataset.augment_counters)

    all_train_labels = [label for label in train_dataset.labels]
    class_weights = compute_class_weight("balanced", classes=np.unique(all_train_labels), y=all_train_labels)
    class_weights = torch.FloatTensor(class_weights).to(torch.device("cuda" if torch.cuda.is_available() else "cpu"))

    sampler = None
    criterion = FocalLoss(alpha=class_weights, gamma=3, reduction="mean")
    device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
    print(f"\nUsing device: {device}")

    model = VisionTransformerModel(num_classes=3).to(device)
    optimizer = torch.optim.Adam(model.parameters(), lr=1e-3, weight_decay=1e-4)
    scheduler = torch.optim.lr_scheduler.ReduceLROnPlateau(optimizer, mode='min', factor=0.1, patience=3, verbose=True)

    for param in model.vit.parameters():
        param.requires_grad = False

    print("\nCreating Dataloaders...")
    train_loader = DataLoader(combined_train_dataset, batch_size=16, shuffle=True)
    validation_loader = DataLoader(validation_dataset, batch_size=16, shuffle=False)
    test_loader = DataLoader(test_dataset, batch_size=16, shuffle=False)

    print("\nStarting model training...")
    trained_model = train_network(
        model,
        optimizer,
        criterion,
        train_loader,
        validation_loader,
        device,
        scheduler,
        logger,
        num_epochs=80,
        early_stopping_patience=5,
        unfreeze_epoch=6
    )

    print("\nEvaluating on the test set:")
    test_loss, test_acc, test_class_accuracies = validate_network(trained_model, test_loader, device, criterion)
    logger.save_test_results(test_loss, test_acc)
    logger.save()
    print("\nTraining and testing completed. All logs saved to Excel.")


if __name__ == "__main__":
    main()
